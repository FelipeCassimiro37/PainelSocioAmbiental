#!/usr/bin/env python3
"""
Gera os arquivos de dados do painel.

Duas fontes possíveis, nesta ordem de prioridade:

  1. Google Sheets  — se a variável de ambiente SHEET_ID estiver definida
                      (é assim que a automação do GitHub roda);
  2. planilha local — o arquivo .xlsx indicado em --arquivo
                      (útil para testar antes de publicar).

Saída, dentro de dados/:
  meta.json          dicionário, lista de municípios e agregados de UF e país
  ind/<id>.json      um vetor com o valor do indicador nos 5.570 municípios
  uf/<sigla>.json    todos os indicadores dos municípios daquela UF

O painel carrega meta.json na abertura e busca os demais sob demanda.
"""
import argparse, csv, hashlib, io, json, os, sys, urllib.parse, urllib.request
from datetime import datetime, timezone
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
SAIDA = RAIZ / "dados"

UF = {
 '11':'RO','12':'AC','13':'AM','14':'RR','15':'PA','16':'AP','17':'TO','21':'MA',
 '22':'PI','23':'CE','24':'RN','25':'PB','26':'PE','27':'AL','28':'SE','29':'BA',
 '31':'MG','32':'ES','33':'RJ','35':'SP','41':'PR','42':'SC','43':'RS','50':'MS',
 '51':'MT','52':'GO','53':'DF',
}
UF_NOME = {
 '11':'Rondônia','12':'Acre','13':'Amazonas','14':'Roraima','15':'Pará','16':'Amapá',
 '17':'Tocantins','21':'Maranhão','22':'Piauí','23':'Ceará','24':'Rio Grande do Norte',
 '25':'Paraíba','26':'Pernambuco','27':'Alagoas','28':'Sergipe','29':'Bahia',
 '31':'Minas Gerais','32':'Espírito Santo','33':'Rio de Janeiro','35':'São Paulo',
 '41':'Paraná','42':'Santa Catarina','43':'Rio Grande do Sul','50':'Mato Grosso do Sul',
 '51':'Mato Grosso','52':'Goiás','53':'Distrito Federal',
}
SIGLA_COD = {v: k for k, v in UF.items()}

TEMAS = {
 "demografia": dict(nome="População",  ramp=["#EFEDF8","#CFCAEB","#AAA3DA","#867CC4","#6459A4","#443B77"]),
 "economia":   dict(nome="Economia",   ramp=["#EAF3F4","#B7DCDC","#78C0C0","#3F9C9F","#1C7178","#0A4A52"]),
 "educacao":   dict(nome="Educação",   ramp=["#FBF1E3","#F3DBB0","#E8BE75","#D69C42","#B4762A","#7E4E17"]),
 "saneamento": dict(nome="Saneamento", ramp=["#EAF1FA","#C0D8F0","#8DB9E2","#5A96CE","#2E71AE","#154C7D"]),
 "seguranca":  dict(nome="Segurança",  ramp=["#F7ECF1","#E5C4D5","#CE95B6","#B06693","#8A4270","#5C2549"]),
}
CASAS = {"int":0,"brl":0,"brl_c":0,"pct":2,"dec1":1,"dec2":2,"dec3":3,"dec4":4}
VAZIOS = {"", "-", "—", "–", "n/a", "na", "nd", "#n/d", "#n/a", "s/d", "null", "nan", "none"}

avisos, erros = [], []


def num(v):
    """Converte célula em número, tolerando 'R$ 1.234,50' e '32,17%'."""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).replace("\xa0", " ").strip()
    if s.lower() in VAZIOS:
        return None
    s = s.replace("R$", "").replace("%", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def arredonda(v, fmt):
    if v is None:
        return None
    n = CASAS.get(fmt, 2)
    v = round(v, n)
    return int(v) if n == 0 else v


# --------------------------------------------------------------- leitura
def do_sheets(sheet_id, aba):
    url = ("https://docs.google.com/spreadsheets/d/" + sheet_id +
           "/gviz/tq?tqx=out:csv&sheet=" + urllib.parse.quote(aba))
    with urllib.request.urlopen(url, timeout=90) as r:
        texto = r.read().decode("utf-8")
    linhas = list(csv.reader(io.StringIO(texto)))
    if not linhas:
        return []
    cab = [c.strip() for c in linhas[0]]
    return [dict(zip(cab, l)) for l in linhas[1:]]


def do_xlsx(caminho, aba):
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        return []
    ws = wb[aba]
    it = ws.iter_rows(values_only=True)
    cab = [str(c).strip() if c is not None else "" for c in next(it)]
    return [dict(zip(cab, linha)) for linha in it]


# ----------------------------------------------------------------- build
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--arquivo", default="base-painel-socioambiental.xlsx")
    args = ap.parse_args()

    sheet_id = os.environ.get("SHEET_ID", "").strip()
    if sheet_id:
        ler, fonte = (lambda aba: do_sheets(sheet_id, aba)), f"Google Sheets ({sheet_id[:12]}…)"
    else:
        caminho = RAIZ / args.arquivo
        if not caminho.exists():
            sys.exit(f"Sem SHEET_ID e sem o arquivo {caminho.name}. Nada a fazer.")
        ler, fonte = (lambda aba: do_xlsx(caminho, aba)), caminho.name

    # dicionário -------------------------------------------------------
    dic_linhas = ler("dicionario")
    if not dic_linhas:
        sys.exit("A aba 'dicionario' está vazia ou não foi encontrada.")
    DIC, IDS = {}, []
    for r in dic_linhas:
        i = str(r.get("id", "")).strip()
        if not i:
            continue
        tema = str(r.get("tema", "")).strip()
        if tema not in TEMAS:
            erros.append(f"dicionario: tema '{tema}' desconhecido em '{i}'")
            continue
        DIC[i] = dict(
            nome=str(r.get("nome", i)).strip(), tema=tema,
            unidade=str(r.get("unidade") or "").strip(),
            formato=str(r.get("formato") or "dec2").strip(),
            somavel=str(r.get("somavel") or "0").strip() in ("1", "sim", "True", "true"),
            ano=str(r.get("ano") or "").strip(),
            fonte=str(r.get("fonte") or "").strip(),
        )
        IDS.append(i)

    # malha: define a ordem canônica dos municípios ---------------------
    malha = json.loads((RAIZ / "malha" / "mun.topo.json").read_text(encoding="utf-8"))
    geoms = malha["objects"][next(iter(malha["objects"]))]["geometries"]
    codigos = [g["properties"]["id"] for g in geoms]
    nomes = [g["properties"]["name"] for g in geoms]
    ufs = [g["properties"]["uf"] for g in geoms]
    pos = {c: i for i, c in enumerate(codigos)}
    n = len(codigos)

    series = {k: [None] * n for k in DIC}

    # municípios, uma aba por UF ---------------------------------------
    com_dado = []
    for cod, sig in sorted(UF.items(), key=lambda kv: kv[1]):
        linhas = ler(sig)
        if not linhas:
            avisos.append(f"aba '{sig}' vazia ou ausente")
            continue
        achou = False
        for r in linhas:
            cid = str(r.get("id") or "").split(".")[0].strip()
            if not cid:
                continue
            if cid not in pos:
                avisos.append(f"{sig}: código '{cid}' não existe na malha")
                continue
            i = pos[cid]
            for k in IDS:
                bruto = r.get(k)
                if bruto is None or str(bruto).strip().lower() in VAZIOS:
                    continue
                if DIC[k]["formato"] == "texto":
                    series[k][i] = str(bruto).strip()
                    achou = True
                    continue
                v = num(bruto)
                if v is None:
                    avisos.append(f"{sig}/{cid}: valor ilegível em '{k}' ({bruto!r})")
                    continue
                series[k][i] = arredonda(v, DIC[k]["formato"])
                achou = True
        if achou:
            com_dado.append(cod)

    # agregados: valor da fonte quando existir, cálculo quando não ------
    oficiais, zeros = {}, {}
    for r in ler("agregados"):
        cod = str(r.get("codigo") or "").strip().upper()
        chave = "BR" if cod in ("BR", "BRASIL", "0") else cod.zfill(2)
        if chave != "BR" and chave not in UF:
            avisos.append(f"agregados: código '{cod}' não reconhecido")
            continue
        vals = {}
        for k in IDS:
            bruto = r.get(k)
            if bruto is None or str(bruto).strip().lower() in VAZIOS:
                continue
            if DIC[k]["formato"] == "texto":
                continue
            v = num(bruto)
            if v is not None:
                vals[k] = arredonda(v, DIC[k]["formato"])
        for k, v in vals.items():
            if v == 0 and not DIC[k]["somavel"]:
                zeros.setdefault(k, []).append(chave)
        if vals:
            oficiais[chave] = vals

    for k, chaves in zeros.items():
        onde = ", ".join("Brasil" if c == "BR" else UF[c] for c in chaves)
        avisos.append(
            f"agregados: '{k}' ({DIC[k]['nome']}) veio zerado em {len(chaves)} recortes "
            f"[{onde}]. Taxa ou percentual zerado costuma ser célula vazia exportada "
            f"como zero — apague essas células na aba 'agregados' e o painel calcula "
            f"a partir dos municípios.")

    POP = series.get("pop_total") or [1] * n
    idx_uf = {}
    for i, u in enumerate(ufs):
        idx_uf.setdefault(u, []).append(i)

    def calcula(k, ids):
        m, s = DIC[k], series[k]
        if m["formato"] == "texto":
            return None
        acc = w = 0.0
        achou = False
        for i in ids:
            v = s[i]
            if v is None:
                continue
            achou = True
            if m["somavel"]:
                acc += v
            else:
                p = POP[i] or 0
                acc += v * p
                w += p
        if not achou:
            return None
        return acc if m["somavel"] else (acc / w if w else None)

    agregados, origem = {}, {}
    for chave, ids in [("BR", list(range(n)))] + [(u, idx_uf.get(u, [])) for u in UF]:
        vals, ofs = {}, []
        for k in IDS:
            of = oficiais.get(chave, {}).get(k)
            if of is not None:
                vals[k] = of
                ofs.append(k)
            else:
                v = calcula(k, ids)
                if v is not None:
                    vals[k] = arredonda(v, DIC[k]["formato"])
        agregados[chave] = vals
        origem[chave] = ofs

    # gravação ---------------------------------------------------------
    for p in (SAIDA / "ind", SAIDA / "uf"):
        p.mkdir(parents=True, exist_ok=True)
    for antigo in list((SAIDA / "ind").glob("*.json")) + list((SAIDA / "uf").glob("*.json")):
        antigo.unlink()

    esc = dict(ensure_ascii=False, separators=(",", ":"))

    # A versão é o resumo do conteúdo, não o relógio: rodar sem mudança de dado
    # gera arquivos idênticos e a automação não cria commit à toa.
    assinatura = hashlib.sha256(
        json.dumps([series, agregados, DIC], sort_keys=True, **esc).encode()
    ).hexdigest()[:12]
    atualizado = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    antigo = SAIDA / "meta.json"
    if antigo.exists():
        try:
            velho = json.loads(antigo.read_text(encoding="utf-8"))
            if velho.get("versao") == assinatura:
                atualizado = velho.get("atualizado", atualizado)
        except Exception:
            pass

    meta = dict(
        versao=assinatura, atualizado=atualizado,
        fonte=fonte, dicionario=DIC, temas=TEMAS,
        ufSiglas=UF, ufNomes=UF_NOME, ufComDado=sorted(com_dado),
        municipios=codigos, nomes=nomes, uf=ufs,
        agregados=agregados, oficiais=origem,
    )
    (SAIDA / "meta.json").write_text(json.dumps(meta, **esc), encoding="utf-8")

    for k in IDS:
        (SAIDA / "ind" / f"{k}.json").write_text(json.dumps(series[k], **esc), encoding="utf-8")

    for cod, ids in idx_uf.items():
        bloco = {k: [series[k][i] for i in ids] for k in IDS}
        (SAIDA / "uf" / f"{UF[cod]}.json").write_text(json.dumps(bloco, **esc), encoding="utf-8")

    # relatório --------------------------------------------------------
    preenchidas = sum(1 for k in series for v in series[k] if v is not None)
    tot = sum(f.stat().st_size for f in SAIDA.rglob("*.json"))
    print(f"Fonte ............. {fonte}")
    print(f"Indicadores ....... {len(IDS)}")
    print(f"Municípios ........ {n}")
    print(f"UFs com dado ...... {len(com_dado)}")
    print(f"Células com dado .. {preenchidas:,}".replace(",", "."))
    print(f"Arquivos gerados .. {1 + len(IDS) + len(idx_uf)}  ({tot/1e6:.2f} MB no total)")
    print(f"meta.json ......... {(SAIDA/'meta.json').stat().st_size/1e3:.0f} KB")
    print(f"Versão ............ {assinatura}  (atualizado em {atualizado})")
    if avisos:
        print(f"\nAVISOS ({len(avisos)}):")
        for a in avisos[:25]:
            print("  ·", a)
        if len(avisos) > 25:
            print(f"  … e mais {len(avisos)-25}")
    if erros:
        print(f"\nERROS ({len(erros)}):")
        for e in erros:
            print("  !", e)
        sys.exit(1)


if __name__ == "__main__":
    main()
