#!/usr/bin/env python3
"""
Gera os dados do Ranking de Competitividade dos Municípios (CLP).

    python scripts/build_clp.py

Lê  fonte/clp-ranking-<edicao>.xlsx  — a planilha publicada pelo CLP, sem
nenhuma edição manual — e escreve:

    dados/clp/meta.json          glossário, dimensões, pilares e pesos (~60 KB)
    dados/clp/ranking.json       os 418 municípios ranqueados, para a lista (~45 KB)
    dados/clp/mun/<CODIGO>.json  o detalhe de um município (~7 KB cada)

O ranking cobre só municípios com 80 mil habitantes ou mais — 418 na 6ª
edição, de 5.570 no país. O painel precisa dizer isso onde não há dado, então
`ranking.json` traz a lista exata dos códigos cobertos.

Armadilhas da planilha, todas tratadas aqui:
  · ws.max_row mente (formatação vazia); os dados vão da linha 4 à 421
  · o cabeçalho tem três níveis, com células mescladas que só têm valor na
    âncora — daí o preenchimento para a direita
  · dado ausente vem como a string '.n', não como célula vazia
  · a variação vem como 'Novo município' nos 14 que entraram nesta edição
  · nomes com espaço no fim ('Nota ', 'Acessos de telefonia móvel ')
  · a caixa das dimensões diverge entre abas (INSTITUIÇÕES × Instituições)
"""
import csv, glob, json, os, re, sys, unicodedata
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    raise SystemExit('Falta a biblioteca openpyxl: pip install openpyxl')

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FONTE = os.path.join(RAIZ, 'fonte')
SAIDA = os.path.join(RAIZ, 'dados', 'clp')

LIN_DADOS = 4          # primeira linha de dados nas abas de resultados
N_IDENT = 21           # colunas A–U: identificação do município
METRICAS = ('Dados brutos', 'Nota normalizada', 'Colocação', 'Delta colocação')

avisos = []


def aviso(msg):
    avisos.append(msg)
    print('  aviso: ' + msg)


def limpa(v):
    """Texto normalizado: sem espaço sobrando, sem quebra de linha."""
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return str(v.year)
    return re.sub(r'\s+', ' ', str(v)).strip()


def chave(v):
    """Comparação insensível a acento, caixa e espaço — para cruzar as abas."""
    t = unicodedata.normalize('NFKD', limpa(v).lower())
    return ''.join(c for c in t if not unicodedata.combining(c))


def numero(v):
    """Valor numérico, ou None quando a planilha marca ausência com '.n'."""
    if v is None or (isinstance(v, str) and limpa(v) in ('.n', '-', '')):
        return None
    try:
        return round(float(v), 6)
    except (TypeError, ValueError):
        return None


def inteiro(v):
    """Colocação: a planilha traz como float (63.0); na tela é 63º."""
    n = numero(v)
    return None if n is None else int(round(n))


def delta(v):
    """Variação de colocação: número, ou None para quem entrou nesta edição."""
    if isinstance(v, str):
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def linhas(ws, ate=None):
    """Só as linhas que têm conteúdo de verdade."""
    for i, row in enumerate(ws.iter_rows(min_row=LIN_DADOS, max_row=ate,
                                         values_only=True), start=LIN_DADOS):
        if any(c is not None for c in row):
            yield i, row


def preenche(valores):
    """Repete para a direita o valor de uma célula mesclada."""
    saida, ultimo = [], None
    for v in valores:
        t = limpa(v)
        if t:
            ultimo = t
        saida.append(ultimo)
    return saida


def cabecalho(ws, nivel):
    return [c.value for c in next(ws.iter_rows(min_row=nivel, max_row=nivel))]


# ---------------------------------------------------------------- glossário
def ler_glossario(wb):
    ws = wb['Glossário de indicadores']
    campos = ['dimensao', 'pilar', 'nome', 'descricao', 'polaridade', 'unidade',
              'observacoes', 'fonte', 'periodo', 'coleta', 'atualizado', 'link', 'novo']
    itens = []
    dim = pil = None
    for row in ws.iter_rows(min_row=3, max_row=67, values_only=True):
        vals = [limpa(v) for v in row[1:14]]
        if not vals[2]:                     # sem nome de indicador: não é linha de dado
            continue
        dim = vals[0] or dim                # dimensão e pilar só na primeira linha do grupo
        pil = vals[1] or pil
        it = dict(zip(campos, [dim, pil] + vals[2:]))
        it['polaridade'] = 'negativa' if chave(it['polaridade']).startswith('nega') else 'positiva'
        it['atualizado'] = chave(it['atualizado']) == 'sim'
        if it['observacoes'] in ('-', ''):
            it['observacoes'] = ''
        itens.append(it)
    if len(itens) != 65:
        aviso('o glossário trouxe %d indicadores; a 6ª edição tem 65' % len(itens))
    return itens


def ler_pesos(wb):
    ws = wb['Pesos']
    dimensao, pilar = {}, {}
    for row in ws.iter_rows(min_row=5, max_row=17, values_only=True):
        d = limpa(row[1])
        if d and chave(d) != 'total' and isinstance(row[4], (int, float)):
            dimensao[chave(d)] = round(float(row[4]), 6)
        pd_, pp = limpa(row[6]), limpa(row[7])
        if pp and chave(pp) != '-' and isinstance(row[9], (int, float)):
            pilar[chave(pp)] = {'dimensao': pd_, 'noRanking': round(float(row[9]), 6),
                                'naDimensao': round(float(row[10]), 6)
                                if isinstance(row[10], (int, float)) else None}
    return dimensao, pilar


# ---------------------------------------------------------------- resultados
def ler_pilares(wb):
    """Notas, colocações e variações por pilar, dimensão e ranking geral."""
    ws = wb['Pilares+dimensões-Detalhamento']
    n1 = preenche(cabecalho(ws, 1))
    n2 = preenche(cabecalho(ws, 2))
    n3 = [limpa(v) for v in cabecalho(ws, 3)]

    blocos = []                       # (rotulo, dimensao, coluna inicial)
    c = N_IDENT
    while c < len(n3):
        if n3[c:c + 3] and chave(n3[c]) == 'nota':
            blocos.append((n2[c], n1[c] if n1[c] and c < len(n1) else None, c))
            c += 3
        else:
            c += 1

    dados = {}
    for _, row in linhas(ws):
        cod = str(row[2])
        d = {}
        for rotulo, _dim, c in blocos:
            d[rotulo] = [numero(row[c]), inteiro(row[c + 1]), delta(row[c + 2])]
        dados[cod] = d
    return blocos, dados


def ler_indicadores(wb):
    """Valor bruto, nota, colocação e variação de cada um dos 65 indicadores."""
    ws = wb['Indicadores - Detalhamento']
    n2 = preenche(cabecalho(ws, 2))
    n3 = [limpa(v) for v in cabecalho(ws, 3)]

    ordem, c = [], N_IDENT
    while c + 3 < len(n3):
        if tuple(n3[c:c + 4]) == METRICAS:
            ordem.append((n2[c], c))
            c += 4
        else:
            c += 1
    if len(ordem) != 65:
        aviso('achei %d blocos de indicador na aba de detalhamento; esperava 65' % len(ordem))

    dados = {}
    for _, row in linhas(ws):
        cod = str(row[2])
        dados[cod] = [[numero(row[c]), numero(row[c + 1]),
                       inteiro(row[c + 2]), delta(row[c + 3])] for _, c in ordem]
    return [nome for nome, _ in ordem], dados


def ler_municipios(wb):
    ws = wb['Pilares+dimensões-Detalhamento']
    campos = [limpa(v) for v in cabecalho(ws, 3)][:N_IDENT]
    quero = {'UF': 'uf', 'Município': 'nome', 'Código IBGE (7)': 'cod', 'Capital': 'capital',
             'G100': 'g100', 'Região': 'regiao', 'População': 'pop',
             'Faixa populacional': 'faixa', 'Latitude': 'lat', 'Longitude': 'lon',
             'Setor econômico predominante no município': 'setor'}
    idx = {campos[i]: i for i in range(len(campos)) if campos[i] in quero}
    out = {}
    for _, row in linhas(ws):
        m = {quero[k]: (limpa(row[i]) if isinstance(row[i], str) else row[i])
             for k, i in idx.items()}
        m['cod'] = str(m['cod'])
        m['capital'] = chave(m.get('capital')) == 'sim'
        m['g100'] = chave(m.get('g100')) == 'sim'
        out[m['cod']] = m
    return out


# ============================================================ leitura do site
# O CLP publica os mesmos resultados de duas maneiras: a planilha, que exige
# preencher um cadastro para baixar, e os arquivos que alimentam o próprio site
# do ranking, que são públicos e abertos. O vigia (scripts/monitor_clp.py) baixa
# os segundos para fonte/clp/, e é daí que estas funções leem.
#
# Conferi as duas fontes uma contra a outra na 6ª edição: os 65 indicadores dos
# 418 municípios batem, com diferença máxima de 0,005 — que é o arredondamento
# do site, que publica com duas casas decimais onde a planilha tem sete. Como o
# painel exibe duas casas, isso não muda um dígito na tela.
#
# O que muda é a forma: no site o valor bruto vem FORMATADO ('58.29%',
# 'R$ 25,923.08'), já multiplicado por 100 e com o símbolo junto. Aqui ele é
# desmontado de volta para número puro, na mesma escala em que a planilha o
# entrega — fração para percentual —, porque é essa escala que o resto do
# programa espera e é ela que a lógica de exibição decide como mostrar.

# Nenhuma unidade é sobreposta aqui, e isso é uma decisão, não um esquecimento.
#
# A tentação foi concreta: entre a 6ª e a 7ª edição o CLP mudou a unidade de
# 'Transparência municipal' de 'Nota normalizada de 0 a 10' para 'Porcentagem',
# enquanto a descrição continuava dizendo 'Nota na Escala Brasil Transparente
# 360'. Parecia engano da fonte, e cheguei a corrigir. Os dados desmentiram: na
# 6ª edição o indicador ia de 2,64 a 10,00, com mediana 6,97; na 7ª vai de 0,00
# a 99,12, com mediana 74,74. O CLP não errou a legenda — reescalou o indicador,
# e a legenda nova é a certa. A correção é que estava errada.
#
# A lição virou regra: mudança de unidade é RELATADA no Pull Request pelo vigia,
# com o número ao lado, e quem decide é quem lê. Um dicionário de exceções aqui
# passaria por cima da fonte com base num palpite meu, em silêncio, para sempre.


def _numero_formatado(texto, unidade=None):
    """
    Devolve o valor na escala da planilha, a partir do texto do site.

    '58.29%'      -> 0.5829   (percentual volta a ser fração)
    'R$ 25,923.08'-> 25923.08 (a vírgula é separador de milhar, à moda inglesa)
    '0.90'        -> 0.90     (sem símbolo, o número já está na escala da planilha)
    ''            -> None     (o site deixa a célula vazia onde a planilha põe
                               '.n'; nos dois casos é ausência de dado)

    Quem manda é o SÍMBOLO no texto, nunca a unidade declarada — e a diferença
    não é teórica. 'Qualidade da informação contábil e fiscal' tem unidade
    'Porcentagem' e o site publica '0.90', a fração crua, sem o símbolo. Dividir
    por 100 porque a unidade diz 'Porcentagem' transformaria 89,84% em 0,90%, nos
    418 municípios de uma vez. Foi o que aconteceu na primeira versão disto, e só
    apareceu porque a saída foi comparada com a da planilha.

    Fora o símbolo, o número fica como a fonte publicou, e isso também é decisão.
    O CLP mudou a formatação entre edições: em 2025 os percentuais saíam como
    '68.81%' e em 2026 saem como '64.64', já multiplicados e sem o símbolo. Não
    há como adivinhar a escala de um valor isolado. Quem resolve isso é a
    detecção que já existe mais abaixo, pela MEDIANA dos 400 e poucos municípios
    daquele indicador: com mediana 0,68 ela multiplica por 100, com mediana 64
    ela não mexe, e a tela mostra 68,81% e 64,64% nos dois casos. Normalizar aqui
    seria um segundo palpite sobre a mesma coisa, e dois palpites discordam.
    """
    s = limpa(texto)
    if not s or s in ('.n', '-', 'null'):
        return None
    porcento = s.endswith('%')
    s = s.replace('R$', '').replace('%', '').replace(',', '').strip()
    try:
        v = float(s)
    except ValueError:
        return None
    return round(v / 100.0 if porcento else v, 6)


def _abre_fonte_site(pasta):
    """(edição, glossário bruto, linhas do CSV) — ou None se não houver."""
    csvs = sorted(glob.glob(os.path.join(pasta, 'ranking-*.csv')))
    if not csvs:
        return None
    caminho = csvs[-1]
    ano = re.search(r'ranking-(\d{4})', os.path.basename(caminho)).group(1)
    gl = os.path.join(pasta, 'glossario-%s.json' % ano)
    if not os.path.exists(gl):
        raise SystemExit('Achei %s mas não o glossário %s ao lado. O vigia grava '
                         'os dois juntos — rode scripts/monitor_clp.py.'
                         % (os.path.basename(caminho), os.path.basename(gl)))
    with open(gl, encoding='utf-8') as f:
        parametros = json.load(f)
    with open(caminho, encoding='utf-8-sig', newline='') as f:
        linhas_csv = list(csv.reader(f))
    return ano, parametros, linhas_csv


def le_do_site(pasta):
    """
    Mesma coisa que os cinco leitores da planilha devolvem, vinda do site.

    Devolve (edição, glossário, pesos, blocos, pilares, nomes, indicadores,
    municípios) — as peças que main() usa, na ordem em que ele as usa.
    """
    ano, par, linhas_csv = _abre_fonte_site(pasta)
    itens = par['pilares_indicadores']
    cab = linhas_csv[0]
    col = {}
    for j, c in enumerate(cab):
        p = c.split('/')
        if len(p) >= 2:
            col[(p[0], p[1])] = j

    # ---- glossário
    gloss = []
    for x in itens:
        if x.get('tipo') != 'Indicador':
            continue
        unidade = limpa(x.get('unidade'))
        obs = limpa(x.get('obs'))
        gloss.append(dict(
            dimensao=limpa(x.get('dimensao')), pilar=limpa(x.get('pilar')),
            nome=limpa(x.get('nome')), descricao=limpa(x.get('desc')),
            polaridade=('negativa' if chave(x.get('polaridade')).startswith('nega')
                        else 'positiva'),
            unidade=unidade, observacoes='' if obs in ('-', '') else obs,
            fonte=limpa(x.get('fonte')), periodo=limpa(x.get('periodo')),
            coleta=limpa(x.get('coleta')),
            atualizado=chave(x.get('atualizados')) == 'sim',
            link=limpa(x.get('link')), novo=limpa(x.get('novo')),
            codigo=limpa(x.get('codigo'))))
    if len(gloss) != 65:
        aviso('o site trouxe %d indicadores; a 6ª edição tinha 65' % len(gloss))

    # ---- pesos. O site publica o peso de cada pilar no ranking; o da dimensão
    # é a soma dos seus pilares e o do pilar dentro da dimensão é a divisão de
    # um pelo outro. Confirmei as duas contas contra a aba 'Pesos' da planilha
    # da 6ª edição: fecham nos treze pilares.
    peso_dim, peso_pil = {}, {}
    for x in itens:
        if x.get('tipo') == 'Pilar' and isinstance(x.get('peso'), (int, float)):
            d = limpa(x.get('dimensao'))
            peso_pil[chave(x.get('nome'))] = {'dimensao': d,
                                              'noRanking': round(float(x['peso']), 6),
                                              'naDimensao': None}
            peso_dim[chave(d)] = round(peso_dim.get(chave(d), 0) + float(x['peso']), 6)
    for k, v in peso_pil.items():
        total = peso_dim.get(chave(v['dimensao']))
        if total:
            v['naDimensao'] = round(v['noRanking'] / total, 6)

    # ---- blocos: os rótulos seguem a convenção que a planilha usa, porque é
    # ela que main() interpreta ('Ranking Geral', 'Dimensão: <nome>', <pilar>).
    blocos = []
    for x in itens:
        tipo, cod, nome = x.get('tipo'), limpa(x.get('codigo')), limpa(x.get('nome'))
        if tipo == 'Geral':
            rot, dim = 'Ranking Geral', None
        elif tipo == 'Dimensão':
            rot, dim = 'Dimensão: %s' % nome, nome
        elif tipo == 'Pilar':
            rot, dim = nome, limpa(x.get('dimensao'))
        else:
            continue
        if (cod, 'dado') in col:
            blocos.append((rot, dim, cod))

    ordem_ind = [g['codigo'] for g in gloss]
    unidade_de = {g['codigo']: g['unidade'] for g in gloss}

    pilares, indicadores = {}, {}
    for linha in linhas_csv[1:]:
        if not linha or not linha[0].strip():
            continue
        cod = linha[0].split('.')[0].strip()
        if not (cod.isdigit() and len(cod) == 7):
            continue

        def campo(codigo, metrica):
            j = col.get((codigo, metrica))
            return linha[j] if j is not None and j < len(linha) else ''

        pilares[cod] = {rot: [numero(campo(c, 'dado')), inteiro(campo(c, 'pos')),
                              delta_texto(campo(c, 'delta'))]
                        for rot, _dim, c in blocos}
        indicadores[cod] = [[_numero_formatado(campo(c, 'bruto'), unidade_de[c]),
                             numero(campo(c, 'dado')), inteiro(campo(c, 'pos')),
                             delta_texto(campo(c, 'delta'))]
                            for c in ordem_ind]

    # ---- municípios
    muns = {}
    for cod, m in par.get('municipios', {}).items():
        cod = str(cod)
        if cod not in pilares:
            continue
        muns[cod] = dict(
            cod=cod, uf=limpa(m.get('UF')), nome=limpa(m.get('nome')),
            capital=chave(m.get('capital')) == 'sim',
            g100=chave(m.get('g100')) == 'sim',
            regiao=limpa(m.get('regiao')), pop=m.get('pop'),
            faixa=limpa(m.get('pop_faixa')), lat=m.get('lat'), lon=m.get('lng'),
            setor=limpa(m.get('setor')))
    faltando = sorted(set(pilares) - set(muns))
    if faltando:
        aviso('%d municípios com resultado mas sem cadastro no glossário: %s'
              % (len(faltando), ', '.join(faltando[:5])))

    nomes_ind = [g['nome'] for g in gloss]
    blocos_pl = [(rot, dim, c) for rot, dim, c in blocos]
    return (ano, gloss, peso_dim, peso_pil, blocos_pl, pilares,
            nomes_ind, indicadores, muns)


def delta_texto(v):
    """
    Variação de colocação a partir do texto do site.

    Na planilha os municípios que entraram nesta edição trazem 'Novo município';
    no site a célula vem vazia. Os dois casos viram None, que o painel mostra
    como 'novo'.
    """
    s = limpa(v)
    if not s or not re.fullmatch(r'[+-]?\d+', s):
        return None
    return int(s)


def grava(caminho, obj):
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    with open(caminho, 'w', encoding='utf-8') as f:
        json.dump(obj, f, separators=(',', ':'), ensure_ascii=False)


def main():
    # Duas fontes possíveis, e a planilha ganha quando existe: ela tem sete
    # casas decimais onde o site tem duas. Na prática o painel exibe duas, mas
    # se algum dia você baixar a planilha oficial e puser em fonte/, ela passa
    # a valer sem precisar mexer em nada — e o site fica como o caminho que o
    # robô consegue percorrer sozinho, sem cadastro nenhum.
    planilhas = sorted(glob.glob(os.path.join(FONTE, 'clp-ranking-*.xlsx')))
    do_site = sorted(glob.glob(os.path.join(FONTE, 'clp', 'ranking-*.csv')))

    edicao_site = None
    if do_site:
        ano_site = re.search(r'ranking-(\d{4})', os.path.basename(do_site[-1])).group(1)
        with open(os.path.join(FONTE, 'clp', 'edicao-%s.json' % ano_site),
                  encoding='utf-8') as f:
            edicao_site = str(json.load(f)['edicao'])

    edicao_planilha = (re.search(r'clp-ranking-(\d+)', os.path.basename(planilhas[-1])).group(1)
                       if planilhas else None)

    # A planilha só ganha se for da MESMA edição ou mais nova. Uma planilha
    # velha esquecida em fonte/ não pode segurar o painel numa edição anterior
    # à que o robô já trouxe — foi exatamente para não depender de ninguém
    # lembrar de apagar arquivo que existe esta comparação.
    usa_site = bool(do_site) and (not planilhas or int(edicao_planilha) < int(edicao_site))

    if usa_site:
        (edicao, gloss, peso_dim, peso_pil, blocos, pilares,
         nomes_ind, indicadores, muns) = le_do_site(os.path.join(FONTE, 'clp'))
        edicao = edicao_site
        print('Lendo os arquivos abertos do site do CLP (%sª edição)' % edicao)
        if planilhas:
            print('  (a planilha %s é da %sª edição, mais antiga — ignorada)'
                  % (os.path.basename(planilhas[-1]), edicao_planilha))
    elif planilhas:
        caminho = planilhas[-1]
        edicao = edicao_planilha
        print('Lendo %s' % os.path.basename(caminho))
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        gloss = ler_glossario(wb)
        peso_dim, peso_pil = ler_pesos(wb)
        blocos, pilares = ler_pilares(wb)
        nomes_ind, indicadores = ler_indicadores(wb)
        muns = ler_municipios(wb)
    else:
        raise SystemExit('Nada para ler: nem fonte/clp-ranking-<edicao>.xlsx, '
                         'nem fonte/clp/ranking-<ano>.csv. Rode '
                         'scripts/monitor_clp.py para buscar do site.')
    print('  %d municípios · %d indicadores · %d blocos de pilar'
          % (len(muns), len(nomes_ind), len(blocos)))

    # --- cruza o glossário com a ordem das colunas de indicador
    porNome = {chave(g['nome']): g for g in gloss}
    ind_meta, sem_gloss = [], []
    for pos, nome in enumerate(nomes_ind):
        g = porNome.get(chave(nome))
        if not g:
            sem_gloss.append(nome)
            g = {'nome': nome, 'dimensao': '', 'pilar': '', 'descricao': '',
                 'polaridade': 'positiva', 'unidade': '', 'fonte': '',
                 'periodo': '', 'observacoes': '', 'atualizado': True}
        # Escala de exibição. O CLP publica os percentuais como fração (0,6399
        # para 63,99%), e "0,64" ao lado da unidade "Porcentagem" confunde.
        # A decisão sai dos dados, não de um palpite: uso a MEDIANA dos 418, não
        # o máximo — indicadores como Endividamento e Crescimento do PIB passam
        # legitimamente de 150% em alguns municípios, e o máximo faria o teste
        # falhar justamente onde ele mais importa.
        un = chave(g['unidade'])
        vals = sorted(abs(d[pos][0]) for d in indicadores.values() if d[pos][0] is not None)
        mediana = vals[len(vals) // 2] if vals else 0
        escala, suf, pre, casas = 1, '', '', 2
        if 'porcent' in un:
            suf = '%'
            if mediana <= 1.5:
                escala = 100
        elif un.startswith('reais'):
            pre, casas = 'R$ ', 2
        elif 'nota normalizada' in un or 'pontuacao' in un:
            casas = 2
        elif 'taxa' in un or 'acessos' in un or 'internacoes' in un:
            casas = 2
        ind_meta.append({'i': pos, 'nome': limpa(g['nome']), 'pilar': limpa(g['pilar']),
                         'dimensao': limpa(g['dimensao']), 'desc': g['descricao'],
                         'pol': g['polaridade'], 'un': g['unidade'], 'fonte': g['fonte'],
                         'periodo': g['periodo'], 'obs': g['observacoes'],
                         'atualizado': g['atualizado'],
                         'esc': escala, 'suf': suf, 'pre': pre, 'casas': casas})
    if sem_gloss:
        aviso('%d indicadores sem entrada no glossário: %s'
              % (len(sem_gloss), ', '.join(sem_gloss[:3])))

    # --- estrutura de exibição: dimensão -> pilares -> indicadores
    porPilar = {}
    for m in ind_meta:
        porPilar.setdefault(chave(m['pilar']), []).append(m['i'])

    geral, dimensoes, pilares_meta = None, [], []
    for rotulo, dim, _c in blocos:
        k = chave(rotulo)
        if k == 'ranking geral':
            geral = rotulo
        elif k.startswith('dimensao:'):
            nome = limpa(rotulo.split(':', 1)[1])
            dimensoes.append({'nome': nome, 'bloco': rotulo,
                              'peso': peso_dim.get(chave(nome))})
        else:
            pilares_meta.append({
                'nome': rotulo, 'bloco': rotulo,
                'dimensao': limpa(dim) if dim else limpa(peso_pil.get(k, {}).get('dimensao')),
                'peso': (peso_pil.get(k) or {}).get('noRanking'),
                'pesoNaDimensao': (peso_pil.get(k) or {}).get('naDimensao'),
                'indicadores': porPilar.get(k, []),
            })
    orfaos = [p['nome'] for p in pilares_meta if not p['indicadores']]
    if orfaos:
        aviso('pilares sem indicador cruzado: %s' % ', '.join(orfaos))

    grava(os.path.join(SAIDA, 'meta.json'), {
        'edicao': int(edicao),
        'fonte': 'CLP — Ranking de Competitividade dos Municípios, %sª edição' % edicao,
        'criterio': 'Municípios com 80 mil habitantes ou mais',
        'geral': geral,
        'dimensoes': dimensoes,
        'pilares': pilares_meta,
        'indicadores': ind_meta,
        'metricas': ['nota', 'posicao', 'variacao'],
    })

    # --- índice da lista: o que a tela inicial precisa, e nada além
    lista = []
    for cod, m in muns.items():
        g = pilares[cod].get(geral, [None, None, None])
        lista.append({'cod': cod, 'nome': m['nome'], 'uf': m['uf'], 'pop': m['pop'],
                      'faixa': m['faixa'], 'regiao': m['regiao'],
                      'capital': m['capital'], 'g100': m['g100'],
                      'nota': g[0], 'pos': g[1], 'var': g[2],
                      'dim': [pilares[cod].get(d['bloco'], [None, None, None])[1]
                              for d in dimensoes]})
    lista.sort(key=lambda x: (x['pos'] is None, x['pos']))
    grava(os.path.join(SAIDA, 'ranking.json'),
          {'edicao': int(edicao), 'municipios': lista})

    # Posições de cada município em cada pilar e em cada indicador, para a lista
    # poder ser ordenada por qualquer critério — não só pelas 3 dimensões. Vai
    # num arquivo à parte (~130 KB) porque só é baixado quando alguém escolhe um
    # critério fora do básico; deixá-lo no ranking.json dobraria o custo de
    # abertura da página para todo mundo.
    cods = [m['cod'] for m in lista]
    grava(os.path.join(SAIDA, 'ordens.json'), {
        'edicao': int(edicao),
        'cods': cods,
        'pilares': [{'nome': p['nome'],
                     'pos': [pilares[c].get(p['bloco'], [None, None, None])[1] for c in cods]}
                    for p in pilares_meta],
        'indicadores': [{'i': m['i'], 'nome': m['nome'], 'pilar': m['pilar'],
                         'dimensao': m['dimensao'],
                         'pos': [indicadores[c][m['i']][2] for c in cods]}
                        for m in ind_meta],
    })

    # Resumo minúsculo (~10 KB) para o Painel Socioambiental saber, na página de
    # um município, se ele é avaliado pelo CLP e em que posição. Sem isso o
    # painel teria de baixar o ranking inteiro só para descobrir isso — ou, pior,
    # ficar calado nos 5.152 municípios que o CLP não avalia.
    grava(os.path.join(SAIDA, 'resumo.json'), {
        'edicao': int(edicao),
        'total': len(lista),
        'fonte': 'CLP — Ranking de Competitividade dos Municípios, %sª edição' % edicao,
        'criterio': 'Municípios com 80 mil habitantes ou mais',
        'mun': {m['cod']: [m['pos'], m['var']] for m in lista},
    })

    # --- um arquivo por município
    total = 0
    for cod, m in muns.items():
        caminho_m = os.path.join(SAIDA, 'mun', cod + '.json')
        grava(caminho_m, {'edicao': int(edicao), 'info': m,
                          'blocos': pilares[cod], 'ind': indicadores[cod]})
        total += os.path.getsize(caminho_m)

    novos = sum(1 for c in muns if pilares[c].get(geral, [None, None, None])[2] is None)
    print('  ranking.json %.0f KB · %d arquivos de município · %.1f MB no total'
          % (os.path.getsize(os.path.join(SAIDA, 'ranking.json')) / 1e3, len(muns), total / 1e6))
    print('  %d municípios entraram nesta edição (sem variação para comparar)' % novos)
    print('\n%d aviso(s).' % len(avisos) if avisos else '\nSem avisos.')


if __name__ == '__main__':
    main()
