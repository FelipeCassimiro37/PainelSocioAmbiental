#!/usr/bin/env python3
"""
Vigia a divulgação mensal do Novo CAGED e atualiza os dados de emprego.

    python scripts/monitor_caged.py            # verifica e, se houver mês novo, grava
    python scripts/monitor_caged.py --forcar   # regrava mesmo sem mês novo
    python scripts/monitor_caged.py --so-checar  # só diz se mudou, não baixa nada

De onde vem o dado
------------------
NÃO do painel Power BI do MTE. Aquele painel é alimentado por uma API interna,
não documentada, que pode mudar sem aviso — e, pior, passar a devolver número
errado em vez de erro. O caminho usado aqui é o arquivo oficial:

  1. a página de divulgação do Novo CAGED redireciona sozinha para o mês mais
     recente (…/novo-caged → …/2026/julho/pagina-inicial), o que entrega o mês
     de referência de graça, numa requisição de alguns KB;
  2. essa página traz o link da pasta do Google Drive com os anexos;
  3. de lá sai o `3-tabelas_<mês>.xlsx`, cuja **Tabela 8.1** já vem agregada
     por município, com estoque, admissões, desligamentos e saldo, e com a
     série inteira desde jan/2020 num único arquivo de ~42 MB.

O recorte de tempo
------------------
Cada indicador tem a natureza dele, e forçar os dois no mesmo período estragava
um dos dois:

  estoque de empregos   é uma fotografia -> vale o mês mais recente publicado
  admissões, desligamentos, saldo   são fluxos -> soma dos ÚLTIMOS 12 MESES

A janela móvel de 12 meses substitui o ano civil. Ano civil tinha dois defeitos:
em setembro o dado já estava com 9 meses de atraso, e a virada do ano fazia o
número dar um salto sem que nada tivesse acontecido no mundo. Com 12 meses
móveis o período tem sempre o mesmo tamanho — logo é sempre comparável — e anda
junto com a fonte. A variação relativa compara o estoque de hoje com o de 12
meses atrás, o mesmo intervalo dos fluxos.

O que é gravado
---------------
  fonte/auto/auto_caged.csv        um município por linha, colunas com o id do
                                   indicador — o build.py já aceita o id como
                                   nome de coluna, então nada muda na planilha
  fonte/auto/auto_caged.meta.json  período de cada indicador e citação da fonte
  fonte/auto/auto_caged.estado.json  último mês visto, para detectar novidade

Salvaguarda
-----------
Antes de gravar, o que calculamos é conferido contra a **Tabela 7.1** da mesma
planilha, que traz Brasil e as 27 UFs na mesma base (com ajustes) mas calculados
pelo MTE por um caminho que não passa pela soma dos municípios. Se a nossa conta
divergir da dele em um único vínculo, em qualquer um dos 28 recortes, nada é
gravado. Desalinhamento de coluna é o modo de falha mais provável desta planilha
(o bloco mensal tem 4 colunas em jan/2020 e 5 nos demais meses) e é silencioso:
sem essa conferência, publicaria número errado com cara de certo.

O "não identificado"
--------------------
A Tabela 8.1 tem uma linha de código 999999: vínculos cujo município não foi
declarado. Ela não é um município e não tem UF, então não entra em nenhum dos
dois — mas **entra no total do Brasil**, porque são vínculos reais e é o Brasil
que serve de comparação nos relatórios do painel. Consequência esperada: o
Brasil não é a soma exata das 27 UFs, exatamente como o MTE publica.
"""
import argparse, csv, io, json, os, re, sys, unicodedata
from datetime import datetime, timezone

import urllib.parse, urllib.request

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAIDA = os.path.join(RAIZ, 'fonte', 'auto')

PAGINA = ('https://www.gov.br/trabalho-e-emprego/pt-br/acesso-a-informacao/'
          'acoes-e-programas/programas-projetos-acoes-obras-e-atividades/'
          'estatisticas-trabalho/novo-caged')

# id do indicador no painel -> rótulo da métrica na Tabela 8.1
METRICAS = {
    'caged_estoque':   'estoque',
    'caged_admissoes': 'admissoes',
    'caged_deslig':    'desligamentos',
    'caged_saldo':     'saldos',
}
# as chaves ficam SEM acento porque tudo passa por chave() antes de ser
# procurado aqui — com 'março' acentuado, o mês de março nunca casaria
MESES = {m: i + 1 for i, m in enumerate(
    'janeiro fevereiro marco abril maio junho julho agosto setembro '
    'outubro novembro dezembro'.split())}

UA = {'User-Agent': 'painel-socioambiental/1.0 (+github actions; dados publicos)'}


def chave(s):
    t = unicodedata.normalize('NFKD', str(s or '').lower())
    return re.sub(r'\s+', ' ', ''.join(c for c in t if not unicodedata.combining(c))).strip()


def busca(url, binario=False, limite=None):
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=180) as r:
        dados = r.read(limite) if limite else r.read()
        return (dados, r.geturl()) if binario else (dados.decode('utf-8', 'replace'), r.geturl())


# ------------------------------------------------------- descobrir o mês novo
def mes_publicado():
    """Mês de referência da divulgação mais recente, via redirecionamento."""
    html, final = busca(PAGINA)
    m = re.search(r'/novo-caged/(\d{4})/([a-zç]+)', final, re.I)
    if not m:
        m = re.search(r'/novo-caged/(\d{4})/([a-zç]+)/pagina-inicial', html, re.I)
    if not m:
        raise SystemExit('Não consegui identificar o mês de divulgação a partir de '
                         + final + '. A página do MTE provavelmente mudou de formato.')
    ano, nome = int(m.group(1)), chave(m.group(2))
    if nome not in MESES:
        raise SystemExit("Mês '%s' não reconhecido na URL %s" % (nome, final))
    return ano, MESES[nome], final


def link_sumario(html, url_mes):
    """
    Endereço do 'Sumário executivo.pdf' daquele mês, para conferência humana.

    Não é usado para calcular nada — vai só na descrição do Pull Request, para
    você abrir e comparar os totais com um documento que não passou por aqui.
    Sai do <a> da própria página, e não de um padrão de nome montado à mão: o
    nome do arquivo muda de mês para mês, e adivinhá-lo daria link quebrado.
    """
    for m in re.finditer(r'href="([^"]+\.pdf)"', html, re.I):
        href = m.group(1)
        if 'sumario' in chave(href):
            return href if href.startswith('http') else urllib.parse.urljoin(url_mes, href)
    return None


def link_planilha(url_mes, html=None):
    """
    Endereço direto da planilha de tabelas, achada pela pasta do Drive.

    A busca é deliberadamente frouxa quanto ao nome. O MTE já publicou esse
    arquivo como '3-tabelas_julho2026.xlsx' e como '3-tabelas_Julho de 2026.xlsx';
    exigir um nome exato quebraria na primeira vez que ele mudasse o padrão de
    novo. O que se exige é: ser .xlsx, ter 'tabela' no nome e ser o maior
    candidato — as tabelas são de longe o maior anexo da divulgação.
    """
    if html is None:
        html, _ = busca(url_mes)
    pasta = re.search(r'drive\.google\.com/drive/folders/([A-Za-z0-9_-]{20,})', html)
    if not pasta:
        raise SystemExit('Não achei a pasta do Google Drive em ' + url_mes +
                         '. O MTE mudou a forma de publicar os anexos.')
    listagem, _ = busca('https://drive.google.com/drive/folders/' + pasta.group(1))

    achados = []
    for fid, nome in re.findall(r'\["([A-Za-z0-9_-]{25,})",(?:[^]]*?)"([^"]*?\.xlsx)"',
                                listagem):
        achados.append((fid, nome))
    for fid, nome in re.findall(
            r'data-id="([A-Za-z0-9_-]{20,})"[^>]*>.*?([^<>"]*?\.xlsx)', listagem, re.S):
        achados.append((fid, nome))

    vistos, candidatos = set(), []
    for fid, nome in achados:
        if fid in vistos:
            continue
        vistos.add(fid)
        if 'tabela' in chave(nome):
            candidatos.append((fid, nome))
    if not candidatos:
        raise SystemExit('A pasta do Drive não trouxe nenhum .xlsx com "tabela" no '
                         'nome. Confira ' + url_mes + ' antes de mexer no script.')
    if len(candidatos) > 1:
        # entre vários, fica o que começa com "3" (é assim que o MTE numera as
        # tabelas desde 2020); se nem isso resolver, o primeiro da lista
        preferido = [c for c in candidatos if chave(c[1]).lstrip().startswith('3')]
        candidatos = preferido or candidatos
        print('  aviso: %d planilhas candidatas na pasta; usando "%s"'
              % (len(vistos), candidatos[0][1]))
    fid, nome = candidatos[0]
    return 'https://drive.google.com/uc?export=download&id=' + fid, nome


# ------------------------------------------------------------- ler a Tabela 8.1
def mapa_codigos():
    """
    A Tabela 8.1 traz o código IBGE de 6 dígitos; o painel usa o de 7. O sétimo
    é o dígito verificador, então a conversão é uma busca na lista de municípios
    que o próprio painel já publica — nada de recalcular dígito à mão.
    """
    caminho = os.path.join(RAIZ, 'dados', 'meta.json')
    if not os.path.exists(caminho):
        raise SystemExit('Preciso de dados/meta.json para converter o código de 6 '
                         'para 7 dígitos. Rode o build.py antes.')
    with open(caminho, encoding='utf-8') as f:
        municipios = json.load(f)['municipios']
    return {str(c)[:6]: str(c) for c in municipios}


def abre(conteudo):
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)


def acha_aba(wb, *prefixos):
    for p in prefixos:
        nome = next((s for s in wb.sheetnames if chave(s).startswith(p)), None)
        if nome:
            return wb[nome]
    raise SystemExit('A planilha não tem a aba %s. Abas: %s'
                     % (' nem '.join(prefixos), wb.sheetnames))


def le_serie(ws, achar_rotulo):
    """
    Lê uma aba de série mensal e devolve {rótulo: {ano_mes: {métrica: valor}}}.

    Serve tanto para a Tabela 8.1 (uma linha por município) quanto para a
    Tabela 7.1 (uma linha por Brasil/região/UF): as duas têm exatamente o mesmo
    cabeçalho de dois níveis, só muda a coluna que identifica a linha.

    O bloco de cada mês NÃO tem largura fixa: jan/2020 tem 4 colunas e os meses
    seguintes têm 5 (ganharam 'Variação Relativa'). Por isso as colunas são
    localizadas pelo par (rótulo do mês, rótulo da métrica) lido do cabeçalho,
    e não por passo fixo — assumir passo fixo desalinha tudo a partir de 2020.

    `achar_rotulo` recebe a linha de cabeçalho dos meses e devolve o índice da
    coluna que identifica cada linha; esse rótulo mora na linha DOS MESES, não
    na das métricas, porque as primeiras colunas ficam vazias na linha de baixo.
    """
    cab_mes = cab_met = None
    dados, cols, ordem_meses = {}, {}, []
    i_rot = None

    for row in ws.iter_rows(values_only=True):
        vals = ['' if v is None else v for v in row]
        txt = [chave(v) for v in vals]

        if cab_met is None:
            # a linha de métricas é a que repete 'estoque' várias vezes
            if txt.count('estoque') >= 2:
                cab_met = txt
                mes_atual = None
                for j, t in enumerate(cab_mes or []):
                    if re.fullmatch(r'[a-z]+/\d{4}', t or ''):
                        mes_atual = t          # começa um bloco mensal
                    elif t:
                        # rótulo que NÃO é mês encerra o bloco. É o que separa os
                        # blocos-resumo do fim ('Acumulado do Ano', 'Últimos 12
                        # Meses') dos meses de verdade: sem isso, as colunas do
                        # resumo entram como se fossem do último mês e o total
                        # nacional sai ~3x maior.
                        mes_atual = None
                    if cab_met[j] in ('estoque', 'admissoes', 'desligamentos', 'saldos') \
                       and mes_atual:
                        nome_mes, ano_mes = mes_atual.split('/')
                        if nome_mes in MESES:
                            k = '%s-%02d' % (ano_mes, MESES[nome_mes])
                            if (k, cab_met[j]) not in cols:   # a primeira vence
                                cols[(k, cab_met[j])] = j
                            if k not in ordem_meses:
                                ordem_meses.append(k)
                    cab_mes[j] = mes_atual or cab_mes[j]
                i_rot = achar_rotulo(cab_mes)
                if i_rot is None:
                    raise SystemExit('Não achei a coluna que identifica as linhas em '
                                     '"%s"' % ws.title)
                continue
            cab_mes = txt[:]        # guarda como possível linha de meses
            continue

        rot = str(vals[i_rot]).split('.')[0].strip()
        if not rot:
            continue                 # rodapés e notas caem aqui
        reg = dados.setdefault(rot, {})
        for (k, met), j in cols.items():
            v = vals[j]
            if isinstance(v, (int, float)):
                reg.setdefault(k, {})[met] = int(round(v))

    if not cols:
        raise SystemExit('Não consegui mapear as colunas de mês de "%s" — o '
                         'cabeçalho da planilha provavelmente mudou.' % ws.title)
    return dados, sorted(ordem_meses)


def le_municipios(wb):
    """
    Tabela 8.1: um município por linha, série com ajustes.

    Devolve (municípios, meses, não_identificado). A linha "999999" não é um
    município — é gente cujo vínculo não traz o município declarado. Ela não
    entra em nenhum município nem em nenhuma UF, mas **entra no total do
    Brasil**, senão o retrato nacional do painel ficaria alguns milhares abaixo
    do que o MTE publica, e é justamente o Brasil que serve de comparação nos
    relatórios.
    """
    ws = acha_aba(wb, 'tabela 8.1', 'tabela 8')
    dados, meses = le_serie(
        ws, lambda cab: next((j for j, t in enumerate(cab)
                              if 'codigo' in t and 'munic' in t), None))

    de6 = mapa_codigos()
    municipios, nao_ident, sobraram = {}, {}, []
    for cod, serie in dados.items():
        if not cod.isdigit() or len(cod) < 6:
            continue
        sete = de6.get(cod[:6])
        if sete:
            municipios[sete] = serie
        elif cod.startswith('999999'):
            nao_ident = serie
        else:
            sobraram.append(cod)
    if sobraram:
        print('  aviso: %d códigos do CAGED não existem na lista do painel (ex.: %s)'
              % (len(sobraram), ', '.join(sobraram[:3])))
    if not nao_ident:
        print('  aviso: a linha "município não identificado" não apareceu; o total '
              'do Brasil pode sair um pouco abaixo do oficial')
    return municipios, meses, nao_ident


def janela_oficial(wb):
    """
    A janela de 12 meses que o MTE declara, lida do cabeçalho da Tabela 1
    ("Últimos 12 Meses** (Ago/25 a Jul/26)").

    Existe para tapar um ponto cego da conferência: comparar os nossos números
    com os da Tabela 7.1 usando a MESMA janela dos dois lados não detecta uma
    janela errada — os dois lados erram junto. Aqui a janela vem de fora, escrita
    pelo próprio MTE, então uma escolha errada de período aparece.

    Devolve (primeiro_mes, ultimo_mes) ou None se o cabeçalho mudar de forma.
    """
    curto = {m[:3]: i for m, i in MESES.items()}
    try:
        ws = acha_aba(wb, 'tabela 1')
    except SystemExit:
        return None
    for n, row in enumerate(ws.iter_rows(values_only=True)):
        if n > 8:
            break
        for v in row:
            m = re.search(r'\(([a-z]{3})/(\d{2,4})\s+a\s+([a-z]{3})/(\d{2,4})\)', chave(v))
            if m and '12 meses' in chave(v):
                def mes(nome, ano):
                    if nome not in curto:
                        return None
                    a = int(ano)
                    return '%04d-%02d' % (a if a > 99 else 2000 + a, curto[nome])
                ini, fim = mes(m.group(1), m.group(2)), mes(m.group(3), m.group(4))
                if ini and fim:
                    return ini, fim
    return None


def le_geografico(wb):
    """
    Tabela 7.1: Brasil, regiões e UFs, na MESMA base da 8.1 (com ajustes).

    É a régua da conferência. Como o MTE calcula esses totais por um caminho
    completamente diferente do nosso (ele não soma os 5.570 municípios), bater
    com ela é prova de que a leitura da 8.1 não desalinhou coluna nenhuma.
    """
    ws = acha_aba(wb, 'tabela 7.1')
    dados, _ = le_serie(
        ws, lambda cab: next((j for j, t in enumerate(cab)
                              if 'regiao' in t or 'uf' in t.split()), None))

    # os rótulos vêm por extenso ("São Paulo"); o painel trabalha com o código
    # de 2 dígitos do IBGE, que dados/meta.json já traz junto com o nome
    caminho = os.path.join(RAIZ, 'dados', 'meta.json')
    with open(caminho, encoding='utf-8') as f:
        nomes_uf = json.load(f)['ufNomes']
    por_nome = {chave(n): c for c, n in nomes_uf.items()}

    oficial = {}
    for rot, serie in dados.items():
        k = chave(rot)
        if k == 'brasil':
            oficial['BR'] = serie
        elif k in por_nome:
            oficial[por_nome[k]] = serie
    if 'BR' not in oficial:
        raise SystemExit('A Tabela 7.1 não trouxe a linha "Brasil" — não dá para '
                         'conferir os totais, então não gravo nada.')
    return oficial


# ------------------------------------------------------------------ o recorte
JANELA = 12            # meses somados nos fluxos


def recorte(meses):
    """
    Dois períodos, um para cada natureza de indicador.

    Estoque é uma fotografia: vale o mês mais recente publicado, e não faz
    sentido envelhecê-lo de propósito.

    Admissões, desligamentos e saldo são fluxos: só significam alguma coisa
    somados sobre um período. A soma é dos **últimos 12 meses**, não do ano
    civil fechado. Ano civil tinha dois defeitos: em setembro o dado já estava
    com 9 meses de atraso, e virar o ano fazia o número dar um salto sem que
    nada tivesse acontecido no mundo. Com 12 meses móveis o período tem sempre
    o mesmo tamanho — logo é sempre comparável — e anda junto com a fonte.
    """
    if not meses:
        raise SystemExit('Nenhum mês encontrado na planilha.')
    ultimo = max(meses)
    janela = sorted(meses)[-JANELA:]
    if len(janela) < JANELA:
        raise SystemExit('A planilha só trouxe %d meses; preciso de %d para a '
                         'janela de fluxos.' % (len(janela), JANELA))
    return ultimo, janela


def rotulo_mes(k):
    nome = [m for m, i in sorted(MESES.items(), key=lambda kv: kv[1])][int(k[5:]) - 1]
    return '%s/%s' % (nome[:3], k[:4])


def acumula(dados, ultimo, janela, meses):
    """Estoque = mês mais recente; fluxos = soma da janela; variação = 12 meses."""
    # a base da variação é o mês ANTERIOR ao primeiro da janela: entre ele e o
    # último mês cabem exatamente os 12 meses somados nos fluxos
    todos = sorted(meses)
    i = todos.index(janela[0])
    anterior = todos[i - 1] if i > 0 else None
    saida = {}
    for cod, serie in dados.items():
        linha = {}
        est = (serie.get(ultimo) or {}).get('estoque')
        if est is not None:
            linha['caged_estoque'] = est
        for ind, met in METRICAS.items():
            if met == 'estoque':
                continue
            vals = [serie[k][met] for k in janela if k in serie and met in serie[k]]
            if vals:
                linha[ind] = sum(vals)
        # variação do estoque no mesmo período dos fluxos: comparo a fotografia
        # de hoje com a de 12 meses atrás. Poderia sair de saldo/estoque, mas aí
        # dependeria de as duas séries estarem na mesma revisão — e o estoque é
        # justamente a série que o MTE mais revisa.
        base = (serie.get(anterior) or {}).get('estoque') if anterior else None
        if est is not None and base:
            linha['caged_var'] = round((est - base) / base * 100, 4)
        if linha:
            saida[cod] = linha
    return saida


def agrega(dados, consolidado, nao_ident, ultimo, janela, meses):
    """
    Soma por UF e para o Brasil.

    Sem isso o painel ficaria incoerente: os municípios viriam do robô e os
    cartões de estado e de Brasil continuariam com os números antigos digitados
    na aba 'agregados' da planilha. O build.py dá preferência ao que a fonte
    declara para UF e país, então basta mandar essas linhas junto.

    O "não identificado" entra só no Brasil. São vínculos reais, que precisam
    contar no retrato nacional; mas como não têm município nem UF declarados,
    não há onde encaixá-los abaixo disso. É por isso que o Brasil daqui NÃO é
    a soma das 27 UFs — e é assim que o MTE também publica.

    A variação relativa NÃO é média das variações municipais — é a variação do
    estoque do próprio agregado, calculada dos dois extremos da janela.
    """
    todos = sorted(meses)
    anterior = todos[todos.index(janela[0]) - 1] if todos.index(janela[0]) > 0 else None
    grupos = {'BR': []}
    for cod in consolidado:
        grupos.setdefault(cod[:2], []).append(cod)
        grupos['BR'].append(cod)

    linhas = {}
    for ch, cods in grupos.items():
        extra = nao_ident if (ch == 'BR' and nao_ident) else None
        linha = {}
        est = sum(consolidado[c]['caged_estoque'] for c in cods
                  if 'caged_estoque' in consolidado[c])
        if extra:
            est += (extra.get(ultimo) or {}).get('estoque', 0)
        linha['caged_estoque'] = est
        for ind, met in METRICAS.items():
            if met == 'estoque':
                continue
            soma = sum(consolidado[c][ind] for c in cods if ind in consolidado[c])
            if extra:
                soma += sum((extra.get(k) or {}).get(met, 0) for k in janela)
            linha[ind] = soma
        if anterior:
            base = sum((dados[c].get(anterior) or {}).get('estoque', 0)
                       for c in cods if c in dados)
            if extra:
                base += (extra.get(anterior) or {}).get('estoque', 0)
            if base and est:
                linha['caged_var'] = round((est - base) / base * 100, 4)
        linhas[ch] = linha
    return linhas


def confere(consolidado, agregados, oficial, ultimo, janela, janela_mte=None):
    """
    Confere o que calculamos contra o número que o próprio MTE publica.

    Antes isto era uma faixa plausível escrita à mão ("estoque nacional entre
    30 e 80 milhões"). Passava em qualquer erro que não fosse gritante. Agora a
    régua é a Tabela 7.1 da mesma planilha: Brasil e as 27 UFs, calculados pelo
    MTE por um caminho que não passa pela soma dos municípios. Se a nossa conta
    e a dele divergirem em um único vínculo, nada é gravado.

    É essa dupla origem que dá sentido à conferência. Comparar a Tabela 8.1 com
    ela mesma não provaria nada.
    """
    problemas = []

    # a janela precisa terminar no mês do estoque e ter os 12 meses seguidos;
    # sem isto, uma janela deslocada passaria batido, porque os dois lados da
    # comparação abaixo usam a mesma janela
    if janela[-1] != ultimo:
        problemas.append('a janela de fluxos termina em %s, mas o estoque é de %s'
                         % (janela[-1], ultimo))
    if len(janela) != JANELA or len(set(janela)) != JANELA:
        problemas.append('a janela tem %d meses; esperava %d' % (len(janela), JANELA))
    if janela_mte and (janela[0], janela[-1]) != janela_mte:
        problemas.append('o MTE declara a janela de 12 meses como %s a %s, e a nossa '
                         'é %s a %s' % (janela_mte[0], janela_mte[1], janela[0], janela[-1]))

    for ch in sorted(agregados):
        nosso = agregados[ch]
        ref = oficial.get(ch)
        onde = 'Brasil' if ch == 'BR' else 'UF %s' % ch
        if not ref:
            problemas.append('%s não aparece na Tabela 7.1' % onde)
            continue
        alvo = {'caged_estoque': (ref.get(ultimo) or {}).get('estoque')}
        for ind, met in METRICAS.items():
            if met == 'estoque':
                continue
            alvo[ind] = sum((ref.get(k) or {}).get(met, 0) for k in janela)
        for ind, esperado in alvo.items():
            obtido = nosso.get(ind)
            if esperado is None:
                problemas.append('%s: a Tabela 7.1 não traz %s' % (onde, ind))
            elif obtido != esperado:
                problemas.append('%s, %s: calculei %s e a Tabela 7.1 do MTE diz %s '
                                 '(diferença de %s)'
                                 % (onde, ind, f'{obtido:,}', f'{esperado:,}',
                                    f'{(obtido or 0) - esperado:,}'))

    br = agregados.get('BR', {})
    adm, des, sal = (br.get('caged_admissoes', 0), br.get('caged_deslig', 0),
                     br.get('caged_saldo', 0))
    if abs((adm - des) - sal) > 5:
        problemas.append('admissões − desligamentos (%s) não bate com o saldo (%s)'
                         % (f'{adm - des:,}', f'{sal:,}'))
    if len(consolidado) < 5000:
        problemas.append('só %d municípios lidos; esperava mais de 5.000' % len(consolidado))

    if problemas:
        raise SystemExit('Conferência falhou, nada foi gravado:\n  - ' +
                         '\n  - '.join(problemas[:12]) +
                         ('\n  … e mais %d' % (len(problemas) - 12) if len(problemas) > 12 else ''))
    return dict(estoque=br.get('caged_estoque', 0), admissoes=adm,
                desligamentos=des, saldo=sal, municipios=len(consolidado),
                conferidoCom='Tabela 7.1 (Brasil + %d UFs)' % (len(agregados) - 1))


# --------------------------------------------------------------------- saída
def estado_atual():
    caminho = os.path.join(SAIDA, 'auto_caged.estado.json')
    if os.path.exists(caminho):
        with open(caminho, encoding='utf-8') as f:
            return json.load(f)
    return {}


def grava(consolidado, agregados, ultimo, janela, totais, mes_pub, nome_arquivo,
          sumario=None):
    os.makedirs(SAIDA, exist_ok=True)
    cols = ['codigo'] + list(METRICAS) + ['caged_var']
    with open(os.path.join(SAIDA, 'auto_caged.csv'), 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(cols)
        # o painel usa o código IBGE de 7 dígitos (a Tabela 8.1 traz 6); as duas
        # últimas famílias de linhas são os agregados: 2 dígitos = UF, 'BR' = país
        for cod in sorted(consolidado) + sorted(k for k in agregados if k != 'BR') + ['BR']:
            linha = consolidado.get(cod) or agregados.get(cod) or {}
            w.writerow([cod] + ['' if linha.get(c) is None else linha[c] for c in cols[1:]])

    # cada indicador leva o rótulo do SEU período: o estoque é uma fotografia de
    # um mês, os fluxos são a soma de 12. Um rótulo só para os dois seria mentira
    # em metade dos casos.
    foto = rotulo_mes(ultimo)
    periodo = '%s a %s' % (rotulo_mes(janela[0]), rotulo_mes(ultimo))
    anos = {ind: (foto if ind == 'caged_estoque' else periodo)
            for ind in list(METRICAS) + ['caged_var']}
    with open(os.path.join(SAIDA, 'auto_caged.meta.json'), 'w', encoding='utf-8') as f:
        json.dump({ind: {'ano': ano,
                         'fonte': 'CAGED — Ministério do Trabalho e Emprego'}
                   for ind, ano in anos.items()},
                  f, ensure_ascii=False, indent=2)

    with open(os.path.join(SAIDA, 'auto_caged.estado.json'), 'w', encoding='utf-8') as f:
        json.dump({'divulgacao': '%04d-%02d' % mes_pub, 'arquivo': nome_arquivo,
                   'sumarioExecutivo': sumario,
                   'mesEstoque': ultimo, 'janelaFluxos': janela,
                   'rotulos': anos,
                   'verificadoEm': datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC'),
                   'totais': totais},
                  f, ensure_ascii=False, indent=2)


def _per(janela):
    return ('%s a %s' % (rotulo_mes(janela[0]), rotulo_mes(janela[-1]))
            if janela and janela[0] != '?' else '?')


def resumo_markdown():
    """Corpo do Pull Request: os totais nacionais, para conferência humana."""
    e = estado_atual()
    if not e:
        return 'Sem estado gravado.'
    t = e.get('totais', {})
    f = lambda n: '{:,}'.format(n).replace(',', '.')
    jan = e.get('janelaFluxos') or ['?']
    linhas = [
        '**Divulgação:** %s  ·  arquivo `%s`' % (e.get('divulgacao'), e.get('arquivo')),
        '', '| indicador | período | Brasil |', '|---|---|---:|',
        '| Estoque de empregos | %s | %s |' % (
            rotulo_mes(e['mesEstoque']) if e.get('mesEstoque') else '?',
            f(t.get('estoque', 0))),
        '| Admissões | %s | %s |' % (_per(jan), f(t.get('admissoes', 0))),
        '| Desligamentos | %s | %s |' % (_per(jan), f(t.get('desligamentos', 0))),
        '| Saldo | %s | %s |' % (_per(jan), f(t.get('saldo', 0))),
        '| Municípios | | %s |' % f(t.get('municipios', 0)),
        '',
        '**Conferência automática:** cada um destes números, mais os de todas as '
        '27 UFs, foi comparado com a %s — a tabela que o próprio MTE calcula por '
        'um caminho que não passa pela soma dos municípios. Bateu sem uma unidade '
        'de diferença; se não tivesse batido, nada teria sido gravado.'
        % t.get('conferidoCom', 'Tabela 7.1'),
        '',
        'O total do Brasil inclui a linha "município não identificado" (vínculos '
        'sem município declarado), por isso o Brasil não é a soma exata das UFs — '
        'é assim que o MTE publica.',
    ]
    if e.get('sumarioExecutivo'):
        linhas += ['',
                   'Para conferir por fora: [Sumário executivo de %s](%s).'
                   % (e.get('divulgacao'), e['sumarioExecutivo'])]
    linhas += ['',
               'Assim que este PR for aceito, o *Atualizar dados do painel* dispara '
               'sozinho e o site passa a mostrar estes números em poucos minutos.']
    return '\n'.join(linhas)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--forcar', action='store_true', help='regrava mesmo sem mês novo')
    ap.add_argument('--so-checar', action='store_true', help='só reporta, não baixa a planilha')
    ap.add_argument('--resumo', action='store_true', help='imprime o corpo do Pull Request')
    args = ap.parse_args()

    if args.resumo:
        print(resumo_markdown())
        return

    ano_p, mes_p, url_mes = mes_publicado()
    atual = '%04d-%02d' % (ano_p, mes_p)
    antes = estado_atual().get('divulgacao')
    print('Divulgação mais recente do MTE: %s' % atual)
    print('Último processado aqui: %s' % (antes or '(nenhum)'))

    novidade = atual != antes
    if os.environ.get('GITHUB_OUTPUT'):
        with open(os.environ['GITHUB_OUTPUT'], 'a') as f:
            f.write('novidade=%s\n' % ('true' if (novidade or args.forcar) else 'false'))
            f.write('divulgacao=%s\n' % atual)

    if args.so_checar:
        print('MUDOU' if novidade else 'sem novidade')
        return
    if not novidade and not args.forcar:
        print('Nada a fazer.')
        return

    html_mes, _ = busca(url_mes)
    sumario = link_sumario(html_mes, url_mes)
    url_xlsx, nome = link_planilha(url_mes, html_mes)
    print('Baixando %s' % nome)
    conteudo, _ = busca(url_xlsx, binario=True)
    print('  %.1f MB' % (len(conteudo) / 1e6))

    wb = abre(conteudo)
    dados, meses, nao_ident = le_municipios(wb)
    oficial = le_geografico(wb)
    jan_mte = janela_oficial(wb)
    print('  %d municípios · %d meses (%s a %s) · régua: Brasil + %d UFs'
          % (len(dados), len(meses), meses[0], meses[-1], len(oficial) - 1))

    ultimo, janela = recorte(meses)
    print('  estoque: %s   ·   fluxos: %s a %s (%d meses)'
          % (rotulo_mes(ultimo), rotulo_mes(janela[0]), rotulo_mes(ultimo), len(janela)))

    consolidado = acumula(dados, ultimo, janela, meses)
    agregados = agrega(dados, consolidado, nao_ident, ultimo, janela, meses)
    totais = confere(consolidado, agregados, oficial, ultimo, janela, jan_mte)
    print('  conferência ok contra a Tabela 7.1 do MTE, sem uma unidade de '
          'diferença em nenhum dos %d recortes' % len(agregados))
    print('  Brasil: estoque %s · admissões %s · desligamentos %s · saldo %s'
          % tuple(f'{totais[k]:,}'.replace(',', '.')
                  for k in ('estoque', 'admissoes', 'desligamentos', 'saldo')))

    grava(consolidado, agregados, ultimo, janela, totais, (ano_p, mes_p), nome,
          sumario)
    print('Gravado em fonte/auto/auto_caged.csv (%d municípios)' % len(consolidado))


if __name__ == '__main__':
    main()
