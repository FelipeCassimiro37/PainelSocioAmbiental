#!/usr/bin/env python3
"""
Gera os arquivos de dados do painel.

FONTE DOS DADOS, nesta ordem de prioridade:
  1. Google Sheets  — se a variável de ambiente SHEET_ID estiver definida
                      (é assim que a automação do GitHub roda);
  2. planilha local — o .xlsx indicado em --arquivo, para testar antes de publicar.

ABAS RECONHECIDAS:
  dicionario      define cada indicador
  agregados       valores próprios de Brasil e das UFs
  AC, AL, … TO    uma aba por estado; o cabeçalho é o id do indicador
  fonte_*         tabela nacional de uma fonte, no formato em que ela publica;
                  o cabeçalho é o nome original da coluna, ligado ao indicador
                  pelo campo 'coluna' do dicionário

As abas fonte_* são lidas depois das abas de estado, então prevalecem em caso
de conflito. O relatório final avisa sempre que isso acontece.

SAÍDA, em dados/:
  meta.json          dicionário, lista de municípios e agregados
  ind/<id>.json      o valor do indicador nos 5.570 municípios
  uf/<SIGLA>.json    todos os indicadores dos municípios daquela UF
"""
import argparse, csv, hashlib, io, json, os, sys, unicodedata
import urllib.parse, urllib.request
from datetime import datetime, timezone
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
SAIDA = RAIZ / "dados"

UF = {
 '11':'RO','12':'AC','13':'AM','14':'RR','15':'PA','16':'AP','17':'TO','21':'MA',
 '22':'PI','23':'CE','24':'RN','25':'PB','26':'PE','27':'AL','28':'SE','29':'BA',
 '31':'MG','32':'ES','33':'RJ','35':'SP','41':'PR','42':'SC','43':'RS','50':'MS',
 '51':'MT','52':'GO','53':'DF',
}
UF_NOME = {
 '11':'Rondônia','12':'Acre','13':'Amazonas','14':'Roraima','15':'Pará','16':'Amapá',
 '17':'Tocantins','21':'Maranhão','22':'Piauí','23':'Ceará','24':'Rio Grande do Norte',
 '25':'Paraíba','26':'Pernambuco','27':'Alagoas','28':'Sergipe','29':'Bahia',
 '31':'Minas Gerais','32':'Espírito Santo','33':'Rio de Janeiro','35':'São Paulo',
 '41':'Paraná','42':'Santa Catarina','43':'Rio Grande do Sul','50':'Mato Grosso do Sul',
 '51':'Mato Grosso','52':'Goiás','53':'Distrito Federal',
}
TEMAS = {
 "demografia": dict(nome="População",  ramp=["#EFEDF8","#CFCAEB","#AAA3DA","#867CC4","#6459A4","#443B77"]),
 "economia":   dict(nome="Economia",   ramp=["#EAF3F4","#B7DCDC","#78C0C0","#3F9C9F","#1C7178","#0A4A52"]),
 "educacao":   dict(nome="Educação",   ramp=["#FBF1E3","#F3DBB0","#E8BE75","#D69C42","#B4762A","#7E4E17"]),
 "saneamento": dict(nome="Saneamento", ramp=["#EAF1FA","#C0D8F0","#8DB9E2","#5A96CE","#2E71AE","#154C7D"]),
 "seguranca":  dict(nome="Segurança",  ramp=["#F7ECF1","#E5C4D5","#CE95B6","#B06693","#8A4270","#5C2549"]),
}
# Casas guardadas no arquivo de dados. Guardamos no mínimo 2 para taxas e
# decimais: o valor da fonte não deve ser perdido na gravação, e a exibição
# decide separadamente quantas casas mostrar.
CASAS = {"int":0,"brl":2,"brl_c":2,"pct":2,"dec1":2,"dec2":2,"dec3":3,"dec4":4}

# Marcas de ausência. As últimas são notação do SIDRA: '-' zero absoluto ou não
# aplicável, '..' não se aplica, '...' valor não disponível, 'X' omitido.
VAZIOS = {"", "-", "—", "–", "n/a", "na", "nd", "#n/d", "#n/a", "s/d", "null",
          "nan", "none", "..", "...", "x"}
CHAVES_ID = {"id", "cod", "cod.", "codigo", "código", "cod ibge", "codigo ibge",
             "código ibge", "cod_ibge", "codigo_ibge", "codigo municipio",
             "codigo do municipio", "municipio_id"}
# colunas descritivas que aparecem nas exportações e não são indicadores
DESCRITIVAS = {"municipio", "município", "nome", "sigla", "uf", "tipo", "estado",
               "brasil e municipio", "brasil e município", "sexo", "cor ou raca",
               "cor ou raça", "nivel", "nível"}

avisos, erros = [], []


def norm(s):
    s = unicodedata.normalize("NFKD", str(s if s is not None else "")).encode("ascii", "ignore").decode()
    return " ".join(s.lower().split())


def num(v):
    """Converte célula em número, tolerando 'R$ 1.234,50' e '32,17%'."""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).replace("\xa0", " ").strip()
    if s.lower() in VAZIOS:
        return None
    s = s.replace("R$", "").replace("%", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def arredonda(v, fmt):
    if v is None:
        return None
    n = CASAS.get(fmt, 2)
    v = round(v, n)
    return int(v) if n == 0 else v


def nivel_de(cod, tipo=""):
    """Descobre se a linha é município, UF ou país."""
    t = norm(tipo)
    if t in ("pais", "brasil", "nacional"):
        return "pais", "BR"
    c = str(cod).split(".")[0].strip()
    if c.upper() in ("BR", "BRASIL") or c in ("0", "00", "1"):
        return "pais", "BR"          # o SIDRA identifica o Brasil como código 1
    if len(c) == 7 and c.isdigit():
        return "mun", c
    if len(c) == 2 and c.isdigit():
        return "uf", c
    if t in ("uf", "estado", "unidade da federacao"):
        return "uf", c.zfill(2)
    return None, None


# ------------------------------------------------------------------- leitura
def _baixa_gviz(sheet_id, aba, extra):
    url = ("https://docs.google.com/spreadsheets/d/" + sheet_id +
           "/gviz/tq?tqx=out:csv&sheet=" + urllib.parse.quote(aba) + extra)
    with urllib.request.urlopen(url, timeout=120) as r:
        return list(csv.reader(io.StringIO(r.read().decode("utf-8"))))


def ler_sheets(sheet_id, aba):
    """Lê uma aba do Google Sheets.

    O gviz tem duas formas de responder. Com 'headers=0' devolve tudo como
    dado, o que é necessário para planilhas com título ou cabeçalho em dois
    níveis; mas quando uma coluna mistura texto no topo e números embaixo ele
    às vezes decide o tipo da coluna e apaga o que não encaixa. Sem o
    parâmetro, ele devolve o cabeçalho como primeira linha, o que é o
    suficiente para tabelas simples. Tentamos as duas e ficamos com a que
    trouxer uma coluna de código reconhecível.
    """
    def nota(linhas):
        """Quanto essa resposta parece um cabeçalho utilizável."""
        if not linhas:
            return -1
        melhor = 0
        for l in linhas[:12]:
            rotulos = sum(1 for c in l if str(c).strip())
            tem_id = any(norm(c) in CHAVES_ID for c in l)
            melhor = max(melhor, rotulos + (1000 if tem_id else 0))
        return melhor

    candidatas, erro = [], None
    for extra in ("&headers=0", ""):
        try:
            candidatas.append(_baixa_gviz(sheet_id, aba, extra))
        except Exception as e:
            erro = e
    if not candidatas:
        if erro is not None:
            avisos.append(f"aba '{aba}': não foi possível ler ({erro})")
        return []
    return max(candidatas, key=nota)


def ler_xlsx(caminho, aba):
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        return []
    return [["" if c is None else c for c in linha]
            for linha in wb[aba].iter_rows(values_only=True)]


def tabela(linhas, rotulo=""):
    """Transforma linhas cruas em (colunas, registros).

    Encontra sozinha a linha de cabeçalho — a primeira que traz uma coluna de
    código — e ignora títulos acima dela. Quando o cabeçalho tem nomes
    repetidos, como nas exportações do SIDRA em que 'Total' e '18 a 24 anos'
    aparecem uma vez por nível de instrução, compõe o nome definitivo com a
    linha de cima preenchida para a direita, virando
    'Superior completo - 18 a 24 anos'.
    """
    if not linhas:
        return [], []
    # O SIDRA repete 'Cód.' em todas as linhas de cabeçalho; vale a última,
    # que é a que traz os nomes das colunas de valor.
    candidatas = [i for i, l in enumerate(linhas[:12])
                  if any(norm(c) in CHAVES_ID for c in l)]
    idx = candidatas[-1] if candidatas else 0
    nomes = [str(c).strip() for c in linhas[idx]]

    uteis = [n for n in nomes if n]
    if len(uteis) != len(set(uteis)) and idx > 0:
        acima, atual, composto = linhas[idx - 1], "", []
        for j, nm in enumerate(nomes):
            c = str(acima[j]).strip() if j < len(acima) else ""
            if c:
                atual = c
            # quando as duas linhas trazem o mesmo rótulo, como nas colunas
            # descritivas do SIDRA, não faz sentido compor "Cód. - Cód."
            composto.append(nm if (not atual or atual == nm) else
                            (f"{atual} - {nm}" if nm else atual))
        nomes = composto
        exemplo = next((x for x in nomes if " - " in x), "")
        avisos.append(f"{rotulo}: cabeçalho em dois níveis, nomes compostos "
                      f"(ex.: '{exemplo}')")

    # colunas sem rótulo recebem um nome interno para continuarem endereçáveis;
    # o Google às vezes devolve o cabeçalho em branco quando a coluna mistura
    # texto no topo com números embaixo
    nomes = [n if n else f"col_{j+1}" for j, n in enumerate(nomes)]

    registros = []
    for l in linhas[idx + 1:]:
        if not any(str(c).strip() for c in l):
            continue
        registros.append({nomes[j]: l[j] for j in range(min(len(nomes), len(l)))})
    return nomes, registros


# --------------------------------------------------------------------- build
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--arquivo", default="base-painel-socioambiental.xlsx")
    args = ap.parse_args()

    sheet_id = os.environ.get("SHEET_ID", "").strip()
    if sheet_id:
        ler = lambda aba: ler_sheets(sheet_id, aba)
        fonte_nome = f"Google Sheets ({sheet_id[:12]}…)"
        listar_abas = None
    else:
        caminho = RAIZ / args.arquivo
        if not caminho.exists():
            sys.exit(f"Sem SHEET_ID e sem o arquivo {caminho.name}. Nada a fazer.")
        ler = lambda aba: ler_xlsx(caminho, aba)
        fonte_nome = caminho.name
        from openpyxl import load_workbook
        listar_abas = load_workbook(caminho, read_only=True).sheetnames

    # ----------------------------------------------------------- dicionário
    _, dic_linhas = tabela(ler("dicionario"), "dicionario")
    if not dic_linhas:
        sys.exit("A aba 'dicionario' está vazia ou não foi encontrada.")

    DIC, IDS, apelidos, abas_declaradas = {}, [], {}, set()
    for r in dic_linhas:
        i = str(r.get("id", "")).strip()
        if not i:
            continue
        tema = str(r.get("tema", "")).strip()
        if tema not in TEMAS:
            erros.append(f"dicionario: tema '{tema}' desconhecido em '{i}'")
            continue
        DIC[i] = dict(
            nome=str(r.get("nome", i)).strip(), tema=tema,
            unidade=str(r.get("unidade") or "").strip(),
            formato=str(r.get("formato") or "dec2").strip(),
            somavel=str(r.get("somavel") or "0").strip() in ("1", "sim", "true", "True"),
            ano=str(r.get("ano") or "").strip(),
            fonte=str(r.get("fonte") or "").strip(),
        )
        # 'grupo' + 'recorte' transformam vários indicadores num item só no
        # painel, com um seletor para a dimensão (faixa etária, sexo, cor…).
        grupo = str(r.get("grupo") or "").strip()
        recorte = str(r.get("recorte") or "").strip()
        if grupo:
            DIC[i]["grupo"] = grupo
            DIC[i]["recorte"] = recorte
        elif recorte:
            avisos.append(f"dicionario: '{i}' tem 'recorte' mas não tem 'grupo' — ignorado")
        IDS.append(i)
        escala = num(r.get("escala")) or 1.0
        for apelido in [i] + str(r.get("coluna") or "").split("|"):
            a = norm(apelido)
            if a:
                apelidos[a] = (i, escala)
        aba = str(r.get("aba") or "").strip()
        if aba:
            abas_declaradas.add(aba)

    # ---------------------------------------------------------------- malha
    malha = json.loads((RAIZ / "malha" / "mun.topo.json").read_text(encoding="utf-8"))
    geoms = malha["objects"][next(iter(malha["objects"]))]["geometries"]
    codigos = [g["properties"]["id"] for g in geoms]
    nomes_mun = [g["properties"]["name"] for g in geoms]
    ufs = [g["properties"]["uf"] for g in geoms]
    pos = {c: i for i, c in enumerate(codigos)}
    n = len(codigos)
    series = {k: [None] * n for k in DIC}
    conflitos = {}

    # ------------------------------------------------------------ leitura
    # Toda aba passa pelo mesmo leitor. A única diferença entre uma aba de
    # estado e uma aba de fonte é o que se espera encontrar nela; a forma de
    # localizar a coluna de código e de casar as colunas com o dicionário é a
    # mesma. Antes as abas de estado exigiam uma coluna chamada exatamente
    # 'id', e quando o Google devolvia esse cabeçalho em branco elas eram
    # puladas linha a linha, sem um único aviso.
    oficiais_fonte = {}

    def processa(aba, tipo, cruas=None):
        cols, linhas = tabela(ler(aba) if cruas is None else cruas, aba)
        if not linhas:
            avisos.append(f"aba '{aba}' vazia ou ausente")
            return None

        col_id = next((c for c in cols if norm(c) in CHAVES_ID), None)
        recurso = ""
        if not col_id:
            col_id = cols[0]
            recurso = f" (sem cabeçalho de código reconhecido; usando a 1ª coluna '{col_id}')"

        mapeadas = {c: apelidos[norm(c)] for c in cols if norm(c) in apelidos}
        if not mapeadas:
            # Aba sem nenhuma coluna de indicador. Numa aba de estado isso é
            # esperado quando a secretaria não publica nada: o Google entrega
            # só as colunas que têm algum valor, e uma aba inteiramente vazia
            # chega com 'id' e 'nome' apenas. Numa aba de fonte é erro de
            # verdade, porque ela existe justamente para trazer indicadores.
            uteis = [c for c in cols if not c.startswith("col_")]
            if tipo == "uf":
                avisos.append(f"{aba}: nenhuma coluna de indicador — o estado "
                              f"provavelmente não publica esses dados")
            else:
                erros.append(f"{aba}: nenhuma coluna casou com o dicionário. "
                             f"Colunas lidas: {uteis[:8]}")
            return None

        lidos = {"mun": 0, "uf": 0, "pais": 0}
        fora = []
        for r in linhas:
            cru = str(r.get(col_id) or "").split(".")[0].strip()
            if not cru:
                continue
            nivel, chave = nivel_de(cru, r.get("tipo", ""))
            if nivel is None:
                if cru.isdigit():
                    avisos.append(f"{aba}: código '{cru}' não reconhecido")
                continue                    # notas de rodapé caem aqui, em silêncio
            if nivel == "mun" and chave not in pos:
                if any(str(r.get(c) or "").strip().lower() not in VAZIOS for c in mapeadas):
                    fora.append(cru)
                continue
            if nivel == "uf" and chave not in UF:
                continue
            lidos[nivel] += 1
            for c, (ind, escala) in mapeadas.items():
                bruto = r.get(c)
                if bruto is None or str(bruto).strip().lower() in VAZIOS:
                    continue
                if DIC[ind]["formato"] == "texto":
                    valor = str(bruto).strip()
                else:
                    v = num(bruto)
                    if v is None:
                        avisos.append(f"{aba}/{cru}: valor ilegível em '{c}' ({bruto!r})")
                        continue
                    valor = arredonda(v * escala, DIC[ind]["formato"])
                if nivel == "mun":
                    i = pos[chave]
                    if series[ind][i] is not None and series[ind][i] != valor:
                        conflitos[(ind, aba)] = conflitos.get((ind, aba), 0) + 1
                    series[ind][i] = valor
                else:
                    oficiais_fonte.setdefault("BR" if nivel == "pais" else chave, {})[ind] = valor

        resumo = (f"  {aba:22} {lidos['mun']:>5} municípios, {lidos['uf']:>2} UFs, "
                  f"{lidos['pais']} país, {len(mapeadas):>2} colunas{recurso}")
        print(resumo)
        if fora:
            avisos.append(f"{aba}: {len(fora)} códigos com dado não existem na malha "
                          f"[{', '.join(fora[:8])}{'…' if len(fora) > 8 else ''}]")
        if not lidos["mun"] and not lidos["uf"] and not lidos["pais"]:
            # linhas existem mas nenhum código foi aproveitado: isso é sempre
            # erro, porque significa que a coluna de código está errada
            erros.append(f"{aba}: {len(linhas)} linhas lidas mas nenhum código "
                         f"aproveitado. Coluna de código usada: '{col_id}'. "
                         f"Primeira linha: {dict(list(linhas[0].items())[:4])}")
        return lidos

    print("Leitura das abas:")
    for cod, sig in sorted(UF.items(), key=lambda kv: kv[1]):
        processa(sig, "uf")

    if listar_abas is None:
        nomes_fonte = sorted(abas_declaradas)   # no Sheets não dá para listar abas
    else:
        nomes_fonte = sorted(set(a for a in listar_abas if a.lower().startswith("fonte_"))
                             | abas_declaradas)
    for aba in nomes_fonte:
        processa(aba, "fonte")

    # ------------------------------------------------------- camada automática
    # Arquivos gerados pelos vigias (scripts/monitor_*.py) e versionados em
    # fonte/auto/. Entram DEPOIS da planilha de propósito: quando um robô já
    # trouxe o dado direto da fonte oficial, ele vale mais que a digitação
    # manual. Cada CSV pode vir com um .meta.json ao lado para atualizar o ano
    # e a citação da fonte dos indicadores que ele preenche — assim o rodapé do
    # painel acompanha o dado sem ninguém precisar editar o dicionário.
    pasta_auto = RAIZ / "fonte" / "auto"
    for csv_auto in sorted(pasta_auto.glob("*.csv")) if pasta_auto.exists() else []:
        import csv as _csv
        with csv_auto.open(encoding="utf-8-sig", newline="") as f:
            amostra = f.read(4096); f.seek(0)
            try:
                dial = _csv.Sniffer().sniff(amostra, delimiters=";,\t")
            except _csv.Error:
                dial = _csv.excel; dial.delimiter = ";"
            cruas = [linha for linha in _csv.reader(f, dial)]
        processa(f"auto/{csv_auto.name}", "fonte", cruas=cruas)

        meta_auto = csv_auto.with_suffix(".meta.json")
        if meta_auto.exists():
            for ind, campos in json.loads(meta_auto.read_text(encoding="utf-8")).items():
                if ind not in DIC:
                    avisos.append(f"{meta_auto.name}: indicador '{ind}' não existe "
                                  f"no dicionário — ano/fonte ignorados")
                    continue
                for campo in ("ano", "fonte", "unidade"):
                    if campos.get(campo):
                        DIC[ind][campo] = str(campos[campo]).strip()

    for (ind, aba), qtd in sorted(conflitos.items()):
        if aba.startswith("auto/"):
            # aqui sobrescrever é o objetivo, não um acidente
            print(f"  '{ind}': {qtd} valores atualizados por {aba}")
        else:
            avisos.append(f"'{ind}': a aba '{aba}' sobrescreveu {qtd} valores vindos "
                          f"das abas de estado")

    com_dado = sorted({ufs[i] for i in range(n) for k in series if series[k][i] is not None})

    # ------------------------------------------------------------ agregados
    oficiais, zeros = {}, {}
    _, agr = tabela(ler("agregados"), "agregados")
    for r in agr:
        cod = str(r.get("codigo") or "").strip().upper()
        chave = "BR" if cod in ("BR", "BRASIL", "0") else cod.zfill(2)
        if chave != "BR" and chave not in UF:
            avisos.append(f"agregados: código '{cod}' não reconhecido")
            continue
        vals = {}
        for k in IDS:
            bruto = r.get(k)
            if bruto is None or str(bruto).strip().lower() in VAZIOS:
                continue
            if DIC[k]["formato"] == "texto":
                continue
            v = num(bruto)
            if v is not None:
                vals[k] = arredonda(v, DIC[k]["formato"])
        for k, v in vals.items():
            if v == 0 and not DIC[k]["somavel"]:
                zeros.setdefault(k, []).append(chave)
        if vals:
            oficiais[chave] = vals
    for chave, vals in oficiais_fonte.items():
        oficiais.setdefault(chave, {}).update(vals)

    idx_uf = {}
    for i, u in enumerate(ufs):
        idx_uf.setdefault(u, []).append(i)

    zerados = {}
    for cod, ids in sorted(idx_uf.items()):
        for k in IDS:
            if DIC[k]["tema"] != "seguranca" or DIC[k]["formato"] == "texto":
                continue
            vistos = [series[k][i] for i in ids if series[k][i] is not None]
            if vistos and not any(v for v in vistos):
                for i in ids:
                    series[k][i] = None
                zerados.setdefault(k, []).append(UF[cod])
    for k, sigs in zerados.items():
        avisos.append(f"'{k}' ({DIC[k]['nome']}) estava zerado em todos os municípios de "
                      f"{len(sigs)} estados [{', '.join(sigs)}] e passou a constar sem dado")
    for k, chaves in zeros.items():
        onde = ", ".join("Brasil" if c == "BR" else UF[c] for c in chaves)
        avisos.append(f"agregados: '{k}' veio zerado em {len(chaves)} recortes [{onde}]. "
                      f"Taxa zerada costuma ser célula vazia exportada como zero — "
                      f"apague a célula e o painel calcula")

    POP = series.get("pop_total") or [1] * n

    def calcula(k, ids):
        m, s = DIC[k], series[k]
        if m["formato"] == "texto":
            return None
        acc = w = 0.0
        achou = False
        for i in ids:
            v = s[i]
            if v is None:
                continue
            achou = True
            if m["somavel"]:
                acc += v
            else:
                p = POP[i] or 0
                acc += v * p
                w += p
        if not achou:
            return None
        return acc if m["somavel"] else (acc / w if w else None)

    agregados, origem = {}, {}
    for chave, ids in [("BR", list(range(n)))] + [(u, idx_uf.get(u, [])) for u in UF]:
        vals, ofs = {}, []
        for k in IDS:
            of = oficiais.get(chave, {}).get(k)
            if of is not None:
                vals[k] = of
                ofs.append(k)
            else:
                v = calcula(k, ids)
                if v is not None:
                    vals[k] = arredonda(v, DIC[k]["formato"])
        agregados[chave] = vals
        origem[chave] = ofs

    # -------------------------------------------------------------- gravação
    for p in (SAIDA / "ind", SAIDA / "uf"):
        p.mkdir(parents=True, exist_ok=True)
    for antigo in list((SAIDA / "ind").glob("*.json")) + list((SAIDA / "uf").glob("*.json")):
        antigo.unlink()

    esc = dict(ensure_ascii=False, separators=(",", ":"))
    assinatura = hashlib.sha256(
        json.dumps([series, agregados, DIC], sort_keys=True, **esc).encode()
    ).hexdigest()[:12]
    atualizado = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    ant = SAIDA / "meta.json"
    if ant.exists():
        try:
            velho = json.loads(ant.read_text(encoding="utf-8"))
            if velho.get("versao") == assinatura:
                atualizado = velho.get("atualizado", atualizado)
        except Exception:
            pass

    meta = dict(versao=assinatura, atualizado=atualizado, fonte=fonte_nome,
                dicionario=DIC, temas=TEMAS, ufSiglas=UF, ufNomes=UF_NOME,
                ufComDado=sorted(com_dado), municipios=codigos, nomes=nomes_mun,
                uf=ufs, agregados=agregados, oficiais=origem)
    (SAIDA / "meta.json").write_text(json.dumps(meta, **esc), encoding="utf-8")
    for k in IDS:
        (SAIDA / "ind" / f"{k}.json").write_text(json.dumps(series[k], **esc), encoding="utf-8")
    for cod, ids in idx_uf.items():
        bloco = {k: [series[k][i] for i in ids] for k in IDS}
        (SAIDA / "uf" / f"{UF[cod]}.json").write_text(json.dumps(bloco, **esc), encoding="utf-8")

    # -------------------------------------------------------------- relatório
    preenchidas = sum(1 for k in series for v in series[k] if v is not None)
    tot = sum(f.stat().st_size for f in SAIDA.rglob("*.json"))
    vazios = [k for k in IDS if not any(v is not None for v in series[k])]
    print(f"\nFonte ............. {fonte_nome}")
    print(f"Indicadores ....... {len(IDS)}")
    print(f"Municípios ........ {n}")
    print(f"UFs com dado ...... {len(com_dado)}")
    print(f"Células com dado .. {preenchidas:,}".replace(",", "."))
    print(f"Arquivos gerados .. {1+len(IDS)+len(idx_uf)}  ({tot/1e6:.2f} MB)")
    print(f"Versão ............ {assinatura}  (atualizado em {atualizado})")
    if vazios:
        print(f"\nSEM NENHUM DADO ({len(vazios)}): {', '.join(vazios)}")
        print("  Estão no dicionário mas não apareceram em nenhuma aba.")
    if avisos:
        print(f"\nAVISOS ({len(avisos)}):")
        for a in avisos[:30]:
            print("  ·", a)
        if len(avisos) > 30:
            print(f"  … e mais {len(avisos)-30}")
    if erros:
        print(f"\nERROS ({len(erros)}):")
        for e in erros:
            print("  !", e)
        sys.exit(1)


if __name__ == "__main__":
    main()
