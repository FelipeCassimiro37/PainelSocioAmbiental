#!/usr/bin/env python3
"""
Vigia as divulgações do SINISA e atualiza os indicadores de saneamento.

    python scripts/monitor_sinisa.py --sondar     # investiga e relata, sem gravar
    python scripts/monitor_sinisa.py --so-checar  # diz se saiu edição nova
    python scripts/monitor_sinisa.py              # baixa e grava, se houver

De onde vem o dado
------------------
Do SINISA — Sistema Nacional de Informações em Saneamento Básico, do Ministério
das Cidades. Ele substituiu o SNIS, que encerrou as atividades em 2023.

A edição é descoberta navegando, nunca por caminho fixo, e há um motivo
concreto: a página `resultados-sinisa` é a edição de 2024 (referência 2023), e
a de 2025 (referência 2024) apareceu como SUBPÁGINA dela,
`resultados-sinisa/resultados-sinisa-2025`. Quem fosse pelo endereço óbvio
pegaria dados um ano mais velhos sem perceber — foi o que aconteceu comigo na
primeira leitura. Então aqui se procura a maior edição publicada.

Três módulos
------------
Resíduos Sólidos, Abastecimento de Água e Esgotamento Sanitário. Cada um tem o
seu pacote na mesma página, e o pacote é achado pelo BLOCO da página em que o
link aparece — nunca pelo nome do arquivo. O motivo cai na cara de quem for pelo
caminho óbvio: o pacote de Água se chama

    SINISA_Resultados_Ref2024.zip

Um nome que promete o resultado inteiro e entrega só a Água. Quem procurasse
'agua' no nome do arquivo não acharia nada; quem tratasse esse arquivo como o
pacote completo leria Água achando que era tudo. Cada bloco da página é ancorado
pela imagem de capa do módulo (CAPA_AGUA_2025.png e companhia), e o link certo é
o rotulado 'Planilhas de Informações e Indicadores' logo abaixo dela.

Base Municipal, não Locais e Regionais
--------------------------------------
Dentro dos pacotes de Água e Esgoto vêm as duas visões. A de Locais e Regionais
é por PRESTADOR e repete o município quando há mais de um; a de Base Municipal
tem uma linha por município, que é o que o painel precisa. Escolher pelo maior
arquivo — como se fazia aqui quando só existia Resíduos — pegaria a errada: no
pacote de Água, a de Locais e Regionais é a maior das duas. A escolha é pelo
caminho dentro do zip, e depois conferida: se algum código de município aparecer
duas vezes, é porque veio a planilha errada, e o script para.

Zero não é o mesmo que ausência
-------------------------------
Este é o cuidado central do script, e em Água e Esgoto ele pesa ainda mais que
em Resíduos, por dois motivos:

  · o SINISA escreve 'Não calculado (condições não atendidas)' — e variantes —
    no lugar do número. É TEXTO no meio de uma coluna numérica: quem converter
    com um float() desatento e cair no except devolvendo 0 acusa o município de
    ter cobertura zero quando o que houve foi falta de condição para calcular.
    Só na coluna de esgoto tratado de 2024 são 1.165 ocorrências;
  · o módulo de Esgoto traz 2.749 municípios dos 5.570. Os outros não têm
    prestador de esgoto declarando — o que NÃO é o mesmo que ter 0% de
    cobertura, ainda que muitos de fato tenham pouco ou nada.

Nos dois casos o município sai com célula vazia, que o painel mostra como 'sem
dado'. Só quem declarou zero recebe zero.

Por isso o CSV traz uma linha para CADA município da malha, e não só para os que
o SINISA cobre: a camada automática do build.py trata célula vazia vinda de
fonte/auto/ como apagamento, e é assim que os zeros indevidos hoje digitados na
planilha do Google saem do ar.
"""
import argparse, csv, io, json, os, re, sys, unicodedata, zipfile
from datetime import datetime, timezone

import urllib.parse, urllib.request

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAIDA = os.path.join(RAIZ, 'fonte', 'auto')
MALHA = os.path.join(RAIZ, 'malha', 'mun.topo.json')

PAGINA = ('https://www.gov.br/cidades/pt-br/acesso-a-informacao/acoes-e-programas/'
          'saneamento/sinisa')

UAS = [
    {'User-Agent': ('Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 '
                    '(KHTML, like Gecko) Chrome/128.0 Safari/537.36'),
     'Accept': '*/*', 'Accept-Encoding': 'identity',
     'Referer': 'https://www.gov.br/cidades/'},
    {'User-Agent': 'painel-socioambiental/1.0 (+github actions; dados publicos)'},
]

FONTE = 'SINISA — Ministério das Cidades'

# ---------------------------------------------------------------- indicadores
# (id no painel, código do SINISA, começo do rótulo na planilha, definição)
#
# O código do SINISA (IAG0001, IES0007…) é o casamento principal: é o
# identificador oficial da variável e não muda quando o texto do rótulo é
# revisado. O rótulo entra como CONFERÊNCIA, não como alternativa — se o código
# estiver numa coluna cujo rótulo não bate, o script para em vez de gravar um
# indicador no lugar de outro.
#
# Resíduos não tem código: naquela planilha a linha que em Água e Esgoto traz os
# códigos traz as unidades ('Percentual', 'Kg/hab.dia'). Lá o casamento é só
# pelo rótulo, como sempre foi — e se compara o COMEÇO do rótulo, que é a parte
# que o SINISA não mexe entre edições.
MODULOS = [
    ('residuos', dict(
        rotulo='Resíduos Sólidos',
        capa='residuos',
        pista='residuo',              # reserva, caso a capa mude de nome
        base_municipal=False,         # o pacote de resíduos não tem essa divisão
        minimo=3000,
        indicadores=[
            ('res_cob_total', None,
             'Cobertura da população total com coleta de resíduos sólidos domiciliares',
             dict(nome='Cobertura de coleta de resíduos', tema='saneamento',
                  unidade='%', formato='pct', somavel='0', faixa=(0, 100))),
            ('res_cob_rural', None,
             'Cobertura da população rural com coleta de resíduos sólidos domiciliares',
             dict(nome='Coleta de resíduos na zona rural', tema='saneamento',
                  unidade='%', formato='pct', somavel='0', faixa=(0, 100))),
            ('res_seletiva', None,
             'Cobertura da população total com coleta seletiva de resíduos sólidos',
             dict(nome='Cobertura de coleta seletiva', tema='saneamento',
                  unidade='%', formato='pct', somavel='0', faixa=(0, 100))),
            ('res_disp_inad', None,
             'Disposição final inadequada de resíduos sólidos urbanos',
             dict(nome='Disposição final inadequada', tema='saneamento',
                  unidade='%', formato='pct', somavel='0', faixa=(0, 100))),
            ('res_massa_pc', None,
             'Massa média per capita de resíduos sólidos urbanos coletados',
             dict(nome='Resíduos coletados por habitante', tema='saneamento',
                  unidade='kg/hab.dia', formato='dec2', somavel='0',
                  faixa=(0, 50), mediana=(0.2, 3.0))),
        ])),
    ('agua', dict(
        rotulo='Abastecimento de Água',
        capa='agua',
        pista='agua',
        base_municipal=True,
        minimo=3000,
        indicadores=[
            ('agua_cob', 'IAG0001',
             'Atendimento da população total com rede de abastecimento de água',
             dict(nome='Cobertura de água', tema='saneamento', unidade='%',
                  formato='pct', somavel='0', faixa=(0, 100))),
            ('perdas', 'IAG2013',
             'Perdas totais de água na distribuição',
             dict(nome='Índice de perdas na distribuição', tema='saneamento',
                  unidade='%', formato='pct', somavel='0', faixa=(0, 100))),
            # Sem teto de propósito: o SINISA publica municípios acima de 1.000
            # l/hab/dia — o maior de 2024 passa de 33.000, num município pequeno
            # com consumo não residencial pesado. São valores oficiais e não cabe
            # a este script censurá-los. O que se exige é que a MEDIANA fique no
            # plausível: é ela que denuncia uma coluna trocada ou uma unidade
            # mudada, que é o risco real.
            ('consumo', 'IAG2006',
             'Consumo total médio per capita de água',
             dict(nome='Consumo médio per capita', tema='saneamento',
                  unidade='l/hab/dia', formato='dec1', somavel='0',
                  faixa=(0, None), mediana=(80, 400))),
            ('emp_agua', 'CAD0005', 'Nome do Prestador',
             dict(nome='Prestador — água', tema='saneamento', unidade='',
                  formato='texto', somavel='0')),
        ])),
    ('esgoto', dict(
        rotulo='Esgotamento Sanitário',
        capa='esgoto',
        pista='esgoto',
        base_municipal=True,
        # Piso mais baixo de propósito: só 2.749 municípios têm prestador de
        # esgoto declarando, e isso é o normal do módulo, não um defeito.
        minimo=2000,
        indicadores=[
            ('esgoto_cob', 'IES0001',
             'Atendimento da população total com rede coletora de esgoto',
             dict(nome='Cobertura de esgoto', tema='saneamento', unidade='%',
                  formato='pct', somavel='0', faixa=(0, 100))),
            # IES0007, e não IES2004 ('Esgoto tratado referido ao esgoto
            # coletado'). A escolha não foi de gosto: comparei os dois com a
            # série que já estava no painel, e IES0007 bateu em 1.459 de 1.459
            # municípios, com diferença mediana de 0,0 ponto, enquanto IES2004
            # errava por 42,7 pontos na mediana. É este o indicador que o painel
            # sempre chamou de 'Esgoto tratado'.
            ('esgoto_trat', 'IES0007',
             'Atendimento dos domicílios totais com coleta e tratamento de esgoto',
             dict(nome='Esgoto tratado', tema='saneamento', unidade='%',
                  formato='pct', somavel='0', faixa=(0, 100))),
            ('emp_esgoto', 'CAD0005', 'Nome do Prestador',
             dict(nome='Prestador — esgoto', tema='saneamento', unidade='',
                  formato='texto', somavel='0')),
        ])),
]


def chave(s):
    t = unicodedata.normalize('NFKD', str(s or '').lower())
    return re.sub(r'\s+', ' ', ''.join(c for c in t if not unicodedata.combining(c))).strip()


def busca(url, cabecalhos=None, ua=None, espera=90, tentativas=4):
    """Requisição com repetição e pausa crescente."""
    import time
    ultimo = None
    for tentativa in range(tentativas):
        if tentativa:
            time.sleep(min(2 ** tentativa, 15))
        h = dict(ua or UAS[min(tentativa, len(UAS) - 1)])
        h.update(cabecalhos or {})
        try:
            return urllib.request.urlopen(
                urllib.request.Request(url, headers=h), timeout=espera)
        except Exception as e:
            ultimo = e
    raise ultimo


def baixa(url, tentativas=6):
    """
    Baixa o arquivo inteiro, retomando de onde parou quando a conexão cai.

    O servidor do Ministério das Cidades corta a transferência no meio com
    alguma frequência (`IncompleteRead`). Como ele aceita `Range`, dá para pedir
    só o pedaço que falta em vez de recomeçar os 19 MB — e, mais importante, o
    tamanho anunciado é conferido no fim: sem isso, um arquivo cortado viraria um
    zip corrompido, ou pior, um zip que abre e vem incompleto.
    """
    import time
    dados = b''
    total = None
    for tentativa in range(tentativas):
        cab = {'Range': 'bytes=%d-' % len(dados)} if dados else {}
        try:
            with busca(url, cab, espera=300, tentativas=2) as r:
                if total is None:
                    n = r.headers.get('Content-Length')
                    faixa = r.headers.get('Content-Range') or ''
                    m = re.search(r'/(\d+)$', faixa)
                    total = int(m.group(1)) if m else (int(n) if n else None)
                while True:
                    pedaco = r.read(1024 * 1024)
                    if not pedaco:
                        break
                    dados += pedaco
            if total is None or len(dados) >= total:
                break
        except Exception:
            if tentativa == tentativas - 1:
                raise
        time.sleep(min(2 ** tentativa, 20))
    if total is not None and len(dados) != total:
        raise SystemExit('O download veio incompleto: %d de %d bytes. Nada foi '
                         'gravado.' % (len(dados), total))
    return dados


# --------------------------------------------------------- achar a edição nova
def edicoes():
    """
    {ano da edição: url da página de resultados}, descoberto navegando.

    Começa na página do SINISA, junta os links de 'resultados-sinisa' e desce um
    nível para achar as subpáginas por ano. A edição sem ano no endereço é a
    primeira (2024); as seguintes vêm com o ano no fim.
    """
    with busca(PAGINA) as r:
        html = r.read().decode('utf-8', 'replace')
    achadas = {}
    vistas = set()
    candidatas = set(re.findall(r'href="([^"]*resultados-sinisa[^"]*)"', html))

    def registra(url):
        m = re.search(r'resultados-sinisa-(20\d{2})/?$', url)
        achadas[int(m.group(1)) if m else 2024] = url.rstrip('/')

    for u in list(candidatas):
        u = urllib.parse.urljoin(PAGINA + '/', u).split('?')[0]
        if '/resultados-sinisa' not in u or u in vistas:
            continue
        vistas.add(u)
        registra(u)
        # desce um nível: a edição nova é subpágina da anterior
        try:
            with busca(u, espera=60, tentativas=2) as r:
                filho = r.read().decode('utf-8', 'replace')
        except Exception:
            continue
        for v in re.findall(r'href="([^"]*resultados-sinisa-20\d{2}[^"]*)"', filho):
            v = urllib.parse.urljoin(u + '/', v).split('?')[0].rstrip('/')
            if v not in vistas:
                vistas.add(v)
                registra(v)
    if not achadas:
        raise SystemExit('Não achei nenhuma página de resultados em ' + PAGINA +
                         '. O Ministério das Cidades mudou o formato do site.')
    return achadas


def pacotes(url_edicao):
    """
    {módulo: url do pacote}, lido pelo BLOCO da página, não pelo nome do arquivo.

    Ver o cabeçalho do arquivo: o pacote de Água se chama
    SINISA_Resultados_Ref2024.zip. Nome nenhum salva quem procurar por 'agua'. O
    que identifica o módulo é a imagem de capa que abre o bloco
    (CAPA_AGUA_2025.png), e o link que interessa é o rotulado 'Planilhas de
    Informações e Indicadores' que vem depois dela.

    A capa é comparada por igualdade exata, e não por 'contém', senão
    CAPA_AGUAS_PLUVIAIS — drenagem, outro módulo — seria confundida com Água.
    """
    with busca(url_edicao) as r:
        html = r.read().decode('utf-8', 'replace')

    marcas = [(m.start(), chave(m.group(1)).replace(' ', '_').strip('_'))
              for m in re.finditer(
                  r'CAPA[_-]([A-Za-z_ÁÉÍÓÚÃÕÇáéíóúãõç]+?)[_-]?20\d{2}\.(?:png|jpe?g)',
                  html, re.I)]

    rotulados, por_nome = {}, {}
    for m in re.finditer(r'<a[^>]+href="([^"]+\.(?:zip|rar|xlsx))"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        url = urllib.parse.urljoin(url_edicao + '/', m.group(1)).split('?')[0]
        rot = chave(re.sub(r'<[^>]+>', ' ', m.group(2)))
        anteriores = [k for p, k in marcas if p < m.start()]
        bloco = anteriores[-1] if anteriores else ''
        if 'planilha' in rot and bloco:
            rotulados.setdefault(bloco, url)
        por_nome.setdefault(chave(url.rsplit('/', 1)[-1]), url)

    saida = {}
    for nome, cfg in MODULOS:
        alvo = rotulados.get(cfg['capa'])
        if not alvo:
            # Reserva: o nome do arquivo. Serve para Resíduos e Esgoto — e é
            # exatamente o que NÃO serve para Água. Se a reserva for usada para
            # Água ela não vai achar nada, e falhar aqui é melhor que adivinhar.
            alvo = next((u for k, u in por_nome.items() if cfg['pista'] in k), None)
        if alvo:
            saida[nome] = alvo
    if not saida:
        raise SystemExit(
            'Não achei nenhum pacote em %s. Capas vistas: %s. Arquivos vistos: %s'
            % (url_edicao, sorted({k for _, k in marcas}), sorted(por_nome)[:8]))
    return saida


# ------------------------------------------------------------ ler a planilha
# Palavras que marcam as visões que o painel NÃO quer: por prestador e agregadas.
DESCARTE = ('locais e regionais', 'locais + regionais', 'uf_mr_br', 'consolidado',
            'balanco', 'regionais')


def acha_planilha(z, cfg):
    """
    A planilha de indicadores do módulo, escolhida pelo caminho dentro do zip.

    Duas exigências: 'indicador' no caminho e nenhuma das palavras de DESCARTE.
    Em Água e Esgoto exige-se ainda 'base municipal', que é a visão de uma linha
    por município. Sem isso, o critério de 'maior arquivo' que servia quando só
    havia Resíduos escolheria a planilha de Locais e Regionais no pacote de
    Água — que é a maior das duas.
    """
    from openpyxl import load_workbook
    candidatos = []
    for info in z.infolist():
        if info.file_size == 0 or not info.filename.lower().endswith(('.xlsx', '.xls')):
            continue
        caminho = chave(info.filename)
        if 'indicador' not in caminho:
            continue
        if any(p in caminho for p in DESCARTE):
            continue
        if cfg['base_municipal'] and 'base municipal' not in caminho:
            continue
        candidatos.append(info)
    if not candidatos:
        raise SystemExit(
            'Nenhuma planilha de indicadores por município no pacote de %s. '
            'Arquivos: %s' % (cfg['rotulo'],
                              [i.filename for i in z.infolist() if i.file_size][:10]))
    escolhido = max(candidatos, key=lambda i: i.file_size)
    wb = load_workbook(io.BytesIO(z.read(escolhido.filename)),
                       read_only=True, data_only=True)
    return escolhido.filename, wb


def valor(v):
    """
    Número, ou None. Nunca zero por desistência.

    No lugar do número, quando não dá para calcular, o SINISA escreve 'Não
    calculado (condições não atendidas)', 'Não calculado (divisão por zero)' ou
    'Não calculado (campos obrigatórios vazios: [gte0019])'. Tudo isso é
    ausência, e ausência não é zero.
    """
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or chave(s).startswith(('nao calculado', 'nao se aplica', 'null',
                                     'nao informado')):
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')   # 1.234,56
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def texto(v):
    s = re.sub(r'\s+', ' ', str(v or '')).strip()
    if not s or chave(s).startswith(('nao calculado', 'null', 'nao informado')):
        return None
    return s


def le_modulo(wb, cfg):
    """
    {codigo_ibge: {id_do_indicador: valor}} e um relatório da leitura.

    O cabeçalho ocupa entre oito e treze linhas e a arrumação MUDA de módulo para
    módulo: em Resíduos os rótulos dos indicadores estão numa linha e as unidades
    na de baixo; em Água e Esgoto há ainda uma linha de códigos entre a de
    unidades e os dados. Em vez de fixar números de linha — que já quebraram uma
    vez — acha-se a coluna do código do IBGE, dali a primeira linha de dados de
    verdade, e procura-se código e rótulo em QUALQUER uma das linhas de cabeçalho
    acima dela.
    """
    ws = wb[wb.sheetnames[0]]
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        raise SystemExit('Planilha de %s vazia.' % cfg['rotulo'])

    i_cod = None
    for linha in linhas[:25]:
        for j, v in enumerate(linha):
            if chave(v) in ('cod_ibge', 'codigo do ibge', 'cod ibge'):
                i_cod = j if i_cod is None else min(i_cod, j)
        if i_cod is not None:
            break
    if i_cod is None:
        raise SystemExit('Não achei a coluna do código do IBGE em %s.' % cfg['rotulo'])

    i_dados = None
    for n, linha in enumerate(linhas):
        c = str(linha[i_cod] or '').split('.')[0].strip()
        if c.isdigit() and len(c) == 7:
            i_dados = n
            break
    if i_dados is None:
        raise SystemExit('Nenhuma linha de município em %s.' % cfg['rotulo'])

    cab = list(range(max(0, i_dados - 5), i_dados))
    largura = max((len(linhas[n]) for n in cab), default=0)

    def celulas(j):
        return [linhas[n][j] for n in cab if j < len(linhas[n])]

    # coluna 'respondeu ao módulo?', quando existe (Resíduos tem)
    i_resp = None
    for n in cab:
        for j, v in enumerate(linhas[n]):
            if chave(v) in ('sim/nao',):
                i_resp = j

    colunas, faltando, trocadas = {}, [], []
    for ident, codigo, rotulo, _ in cfg['indicadores']:
        alvo = chave(rotulo)[:55]
        achou = None
        if codigo:
            for j in range(largura):
                if any(chave(c) == chave(codigo) for c in celulas(j)):
                    achou = j
                    break
            if achou is not None and not any(chave(c).startswith(alvo)
                                             for c in celulas(achou)):
                # O código existe, mas numa coluna cujo rótulo não corresponde:
                # ou o SINISA renumerou, ou trocou o texto. Parar é o certo —
                # gravar seria pôr um indicador no lugar de outro em silêncio.
                trocadas.append('%s (%s): o rótulo dessa coluna é %r'
                                % (ident, codigo,
                                   next((str(c)[:60] for c in celulas(achou) if c), '')))
                achou = None
        if achou is None:
            for j in range(largura):
                if any(chave(c).startswith(alvo) for c in celulas(j)):
                    achou = j
                    break
        if achou is None:
            faltando.append('%s (%s)' % (ident, codigo or rotulo[:40]))
        else:
            colunas[ident] = achou
    if trocadas:
        raise SystemExit(
            'No módulo de %s, o código do SINISA está numa coluna cujo rótulo não '
            'corresponde:\n  - %s\nNada foi gravado. Confira a planilha antes de '
            'mexer no script.' % (cfg['rotulo'], '\n  - '.join(trocadas)))
    if faltando:
        raise SystemExit(
            'A planilha de %s não traz estes indicadores: %s. Os códigos ou os '
            'rótulos mudaram — confira antes de mexer no script.'
            % (cfg['rotulo'], ', '.join(faltando)))

    formatos = {i: d['formato'] for i, _, _, d in cfg['indicadores']}
    dados, sem_resposta, repetidos = {}, 0, []
    for linha in linhas[i_dados:]:
        cod = str(linha[i_cod] or '').split('.')[0].strip()
        if not (cod.isdigit() and len(cod) == 7):
            continue
        if i_resp is not None and not chave(linha[i_resp]).startswith('sim'):
            sem_resposta += 1
            continue                       # NÃO vira zero: fica sem dado
        if cod in dados:
            repetidos.append(cod)
            continue
        reg = {}
        for ident, j in colunas.items():
            if j >= len(linha):
                continue
            if formatos[ident] == 'texto':
                v = texto(linha[j])
            else:
                v = valor(linha[j])
                v = None if v is None else round(v, 4)
            if v is not None:
                reg[ident] = v
        dados[cod] = reg
    if repetidos:
        raise SystemExit(
            'A planilha de %s repete %d códigos de município (ex.: %s). Isso é a '
            'marca da visão por prestador — veio a planilha errada do pacote.'
            % (cfg['rotulo'], len(repetidos), ', '.join(repetidos[:5])))
    return dados, dict(municipios=len(dados), sem_resposta=sem_resposta)


# ------------------------------------------------------------------ conferir
def confere(cfg, dados, relatorio):
    """
    Barreiras antes de gravar.

    Sem uma régua externa como a do CAGED, o que dá para exigir é coerência:
    percentual entre 0 e 100, mediana no plausível onde a grandeza tem ordem
    conhecida, e uma quantidade de municípios compatível com a coleta do módulo.
    """
    problemas = []
    if len(dados) < cfg['minimo']:
        problemas.append('%s: só %d municípios na planilha; esperava mais de %d'
                         % (cfg['rotulo'], len(dados), cfg['minimo']))

    resumo = dict(relatorio)
    for ident, _, _, d in cfg['indicadores']:
        if d['formato'] == 'texto':
            vals = [r[ident] for r in dados.values() if r.get(ident)]
            if vals:
                resumo[ident] = dict(com_dado=len(vals), distintos=len(set(vals)))
            continue
        vals = sorted(r[ident] for r in dados.values() if ident in r)
        if not vals:
            problemas.append('%s: nenhum valor em %s' % (cfg['rotulo'], ident))
            continue
        piso, teto = d.get('faixa', (None, None))
        fora = [v for v in vals
                if (piso is not None and v < piso - 0.01)
                or (teto is not None and v > teto + 0.01)]
        if fora:
            problemas.append('%s/%s: %d valores fora de %s–%s (ex.: %s)'
                             % (cfg['rotulo'], ident, len(fora), piso,
                                teto if teto is not None else 'sem teto', fora[:3]))
        mediana = vals[len(vals) // 2]
        if d.get('mediana'):
            lo, hi = d['mediana']
            if not (lo <= mediana <= hi):
                problemas.append('%s/%s: mediana de %.2f fora do plausível (%s a %s)'
                                 % (cfg['rotulo'], ident, mediana, lo, hi))
        resumo[ident] = dict(com_dado=len(vals), mediana=round(mediana, 2),
                             zeros=sum(1 for v in vals if v == 0),
                             maximo=round(vals[-1], 2))
    if problemas:
        raise SystemExit('Conferência falhou, nada foi gravado:\n  - ' +
                         '\n  - '.join(problemas))
    return resumo


def municipios_da_malha():
    """Os códigos dos municípios do painel, na mesma fonte que o build.py usa."""
    with open(MALHA, encoding='utf-8') as f:
        malha = json.load(f)
    geoms = malha['objects'][next(iter(malha['objects']))]['geometries']
    return [g['properties']['id'] for g in geoms]


def todos_os_indicadores():
    for nome, cfg in MODULOS:
        for ident, codigo, rotulo, d in cfg['indicadores']:
            yield nome, cfg, ident, codigo, rotulo, d


def comparacao_com_o_painel(dados):
    """
    Quanto muda no que já está no ar — para o Pull Request contar, não para
    barrar. O painel guarda os valores em dados/ind/<indicador>.json, na ordem
    dos municípios de dados/meta.json.
    """
    meta = os.path.join(RAIZ, 'dados', 'meta.json')
    if not os.path.exists(meta):
        return {}
    with open(meta, encoding='utf-8') as f:
        cods = [str(c) for c in json.load(f)['municipios']]
    saida = {}
    for _, _, ident, _, _, d in todos_os_indicadores():
        caminho = os.path.join(RAIZ, 'dados', 'ind', ident + '.json')
        if not os.path.exists(caminho):
            continue
        with open(caminho, encoding='utf-8') as f:
            serie = json.load(f)
        antes = dict(zip(cods, serie))
        igual = muda = entra = zerado = perde = 0
        for c in cods:
            a = antes.get(c)
            b = dados.get(c, {}).get(ident)
            if d['formato'] != 'texto':
                a = None if a is None else round(float(a), 2)
                b = None if b is None else round(float(b), 2)
            vazio_antes = a is None or a == ''
            vazio_agora = b is None or b == ''
            if vazio_agora:
                # Duas coisas muito diferentes moram aqui, e juntá-las esconderia
                # justamente o que este trabalho conserta. Um zero que vira "sem
                # dado" é o conserto: era um município que o SINISA não cobre e
                # que a digitação na planilha marcou como zero. Um valor de
                # verdade que vira "sem dado" é perda, e merece ser olhada.
                if vazio_antes:
                    continue
                elif a == 0:
                    zerado += 1
                else:
                    perde += 1
            elif vazio_antes:
                entra += 1
            elif a == b or (d['formato'] != 'texto'
                            and abs(float(a) - float(b)) < 0.02):
                # Inclui o zero que continua zero. Em Resíduos são milhares:
                # 3.439 municípios declararam 0% de coleta seletiva, e isso é
                # informação, não ausência. Contá-los como novidade encheria o
                # relatório de movimento que não existe.
                igual += 1
            elif a == 0:
                # Era zero e agora tem número: o zero da planilha estava no
                # lugar de um dado que a fonte oficial tem.
                entra += 1
            else:
                muda += 1
        saida[ident] = dict(igual=igual, muda=muda, entra=entra,
                            zerado=zerado, perde=perde)
    return saida


# --------------------------------------------------------------------- saída
def estado_atual():
    caminho = os.path.join(SAIDA, 'auto_sinisa.estado.json')
    if os.path.exists(caminho):
        with open(caminho, encoding='utf-8') as f:
            return json.load(f)
    return {}


def grava(dados, edicao, referencia, totais, origens):
    os.makedirs(SAIDA, exist_ok=True)
    ids = [ident for _, _, ident, _, _, _ in todos_os_indicadores()]

    # Uma linha por município da malha, e não só pelos que o SINISA cobre: é o
    # que permite ao build.py apagar os zeros indevidos que estão na planilha.
    with open(os.path.join(SAIDA, 'auto_sinisa.csv'), 'w',
              encoding='utf-8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['codigo'] + ids)
        for cod in municipios_da_malha():
            reg = dados.get(cod, {})
            w.writerow([cod] + ['' if reg.get(i) is None else reg[i] for i in ids])

    meta = {}
    for _, _, ident, _, _, d in todos_os_indicadores():
        meta[ident] = {k: v for k, v in d.items() if k not in ('faixa', 'mediana')}
        meta[ident].update(ano=str(referencia), fonte=FONTE)
    with open(os.path.join(SAIDA, 'auto_sinisa.meta.json'), 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    with open(os.path.join(SAIDA, 'auto_sinisa.estado.json'), 'w', encoding='utf-8') as f:
        json.dump({'edicao': edicao, 'referencia': referencia, 'origens': origens,
                   'verificadoEm': datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC'),
                   'totais': totais,
                   'comparacao': comparacao_com_o_painel(dados)},
                  f, ensure_ascii=False, indent=2)


def resumo_markdown():
    e = estado_atual()
    if not e:
        return 'Sem estado gravado.'
    t = e.get('totais', {})
    comp = e.get('comparacao', {})
    f = lambda n: '{:,}'.format(n).replace(',', '.')
    linhas = ['**SINISA %s** · ano de referência %s'
              % (e.get('edicao'), e.get('referencia')), '']
    for nome, cfg in MODULOS:
        bloco = t.get(nome) or {}
        if not bloco:
            continue
        linhas += ['### %s' % cfg['rotulo'], '']
        linhas.append(
            '%s municípios na planilha%s.'
            % (f(bloco.get('municipios', 0)),
               ('; %s não responderam ao módulo e ficam **sem dado**, não como zero'
                % f(bloco['sem_resposta'])) if bloco.get('sem_resposta') else ''))
        linhas += ['', '| indicador | com dado | mediana | em zero | entram | '
                   'mudam | zeros corrigidos | perdem |',
                   '|---|---:|---:|---:|---:|---:|---:|---:|']
        for ident, _, _, d in cfg['indicadores']:
            b = bloco.get(ident)
            if not b:
                continue
            c = comp.get(ident, {})
            linhas.append('| %s | %s | %s | %s | %s | %s | %s | %s |'
                          % (d['nome'], f(b['com_dado']), b.get('mediana', '—'),
                             f(b.get('zeros', 0)), f(c.get('entra', 0)),
                             f(c.get('muda', 0)), f(c.get('zerado', 0)),
                             f(c.get('perde', 0))))
        linhas.append('')
    linhas += [
        '**As quatro últimas colunas**, comparando com o que está no ar agora. '
        '*Entram*: estavam sem valor e passam a ter número. *Mudam*: tinham valor '
        'e ele muda. *Zeros corrigidos*: estavam como **0** e passam a **sem '
        'dado** — municípios que o SINISA não cobre e que a digitação na planilha '
        'marcou com zero; no mapa, "tem cobertura e ela é zero" vira "não se '
        'sabe", que é o que os dados de fato dizem. *Perdem*: tinham um valor de '
        'verdade e ficam sem dado — esta é a coluna a olhar com desconfiança, e '
        'ela deve ficar em zero.',
        '',
        'Assim que este Pull Request for aceito, o *Atualizar dados do painel* '
        'dispara sozinho e os valores entram no mapa.',
    ]
    return '\n'.join(linhas)


# ------------------------------------------------------------------ processo
def processa(url_edicao, so=None):
    """Baixa e lê os módulos. Devolve (dados unidos, totais, origens)."""
    urls = pacotes(url_edicao)
    unidos, totais, origens = {}, {}, {}
    for nome, cfg in MODULOS:
        if so and nome != so:
            continue
        if nome not in urls:
            raise SystemExit('A página %s não traz o pacote de %s. Pacotes achados: %s'
                             % (url_edicao, cfg['rotulo'], sorted(urls)))
        url = urls[nome]
        print('  %s: %s' % (cfg['rotulo'], url.rsplit('/', 1)[-1]))
        conteudo = baixa(url)
        print('    %.1f MB' % (len(conteudo) / 1e6))
        z = zipfile.ZipFile(io.BytesIO(conteudo))
        arquivo, wb = acha_planilha(z, cfg)
        print('    planilha: %s' % arquivo.rsplit('/', 1)[-1])
        dados, rel = le_modulo(wb, cfg)
        totais[nome] = confere(cfg, dados, rel)
        origens[nome] = url
        for cod, reg in dados.items():
            unidos.setdefault(cod, {}).update(reg)
        print('    %d municípios · %d indicadores'
              % (len(dados), len(cfg['indicadores'])))
    return unidos, totais, origens


def sondar(so=None):
    eds = edicoes()
    print('Edições publicadas: %s' % ', '.join(str(a) for a in sorted(eds)))
    ed = max(eds)
    print('Mais recente: SINISA %d (referência %d)' % (ed, ed - 1))
    print('  %s' % eds[ed])
    unidos, totais, _ = processa(eds[ed], so)
    print()
    for nome, cfg in MODULOS:
        if nome not in totais:
            continue
        print('%s:' % cfg['rotulo'])
        for ident, _, _, d in cfg['indicadores']:
            b = totais[nome].get(ident)
            if not b:
                continue
            if d['formato'] == 'texto':
                print('  %-34s %5d com dado · %d nomes distintos'
                      % (d['nome'][:34], b['com_dado'], b['distintos']))
            else:
                print('  %-34s %5d com dado · mediana %8.2f · %4d em zero · máx %.2f'
                      % (d['nome'][:34], b['com_dado'], b['mediana'], b['zeros'],
                         b['maximo']))
    print()
    print('Municípios com algum dado de saneamento: %d' % len(unidos))
    comp = comparacao_com_o_painel(unidos)
    if comp:
        print('Contra o que já está no painel:')
        for ident, c in comp.items():
            print('  %-12s iguais=%5d  mudam=%4d  entram=%4d  saem=%4d'
                  % (ident, c['igual'], c['muda'], c['entra'], c['sai']))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sondar', action='store_true')
    ap.add_argument('--so-checar', action='store_true')
    ap.add_argument('--forcar', action='store_true')
    ap.add_argument('--resumo', action='store_true')
    ap.add_argument('--modulo', choices=[n for n, _ in MODULOS],
                    help='trabalhar só um módulo (para investigar)')
    args = ap.parse_args()

    if args.resumo:
        print(resumo_markdown())
        return

    try:
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:
        pass

    if args.sondar:
        sondar(args.modulo)
        return

    eds = edicoes()
    ed = max(eds)
    antes = estado_atual()
    # Muda a edição OU muda a lista de módulos lidos. A segunda metade não é
    # luxo: quando este script passou a ler Água e Esgoto, a edição continuou
    # sendo a mesma de antes, e sem isto a rotina diria 'nada a fazer' para
    # sempre — só rodaria com --forcar, e ninguém lembraria.
    modulos_agora = sorted(n for n, _ in MODULOS)
    lidos_antes = sorted(antes.get('origens') or {})
    novidade = ed != antes.get('edicao') or modulos_agora != lidos_antes
    print('Edição mais recente do SINISA: %d' % ed)
    print('Último processado aqui: %s (módulos: %s)'
          % (antes.get('edicao') or '(nenhum)', ', '.join(lidos_antes) or '(nenhum)'))
    if os.environ.get('GITHUB_OUTPUT'):
        with open(os.environ['GITHUB_OUTPUT'], 'a') as f:
            f.write('novidade=%s\n' % ('true' if (novidade or args.forcar) else 'false'))
            f.write('edicao=%d\n' % ed)
    if args.so_checar:
        print('MUDOU' if novidade else 'sem novidade')
        return
    if not novidade and not args.forcar:
        print('Nada a fazer.')
        return

    unidos, totais, origens = processa(eds[ed], args.modulo)
    grava(unidos, ed, ed - 1, totais, origens)
    print('Gravado em fonte/auto/auto_sinisa.csv (%d municípios com dado)' % len(unidos))


if __name__ == '__main__':
    main()
